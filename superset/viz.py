# -*- coding: utf-8 -*-
"""This module contains the 'Viz' objects

These objects represent the backend of all the visualizations that
Superset can render.
"""
from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals

import codecs
import copy
import hashlib
import inspect
import logging
import math
import re
import traceback
import uuid
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from itertools import product
from functools import lru_cache

import geohash
import numpy as np
import pandas as pd
import polyline
import simplejson as json
from dateutil import relativedelta as rdelta
from flask import escape, request, g
from flask_babel import gettext as __
from flask_babel import lazy_gettext as _
from geopy.point import Point
from markdown import markdown
from openpyxl.styles.borders import BORDER_THIN, Border
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Side
from pandas.tseries.frequencies import to_offset
from pandas.api.types import is_datetime64_any_dtype
from six import string_types
from six.moves import cPickle as pkl, reduce

from sqlalchemy import func, Float, ARRAY, String, text, case, column, Text
from superset import app, cache, get_css_manifest_files, utils
from superset.formatters import ExtendedHTMLFormatter
from superset.utils import DTTM_ALIAS, JS_MAX_INTEGER, merge_extra_filters, merge_where

config = app.config
stats_logger = config.get('STATS_LOGGER')

SUB_TOTALS_DEFAULT_AGGR_FUNC = 'sum'
YANDEX_ICONS = config.get('YANDEX_MAP_ICON_PATH')


class BaseViz(object):
    """All visualizations derive this base class"""

    viz_type = None
    verbose_name = 'Base Viz'
    credits = ''
    is_timeseries = False
    default_fillna = 0
    cache_type = 'df'

    def __init__(self, datasource, form_data, force=False):
        if not datasource:
            raise Exception(_('Viz is missing a datasource'))
        self.datasource = datasource
        self.request = request
        self.viz_type = form_data.get('viz_type')
        self.form_data = form_data
        self.query = ''
        self.total_found = 0
        self.token = self.form_data.get(
            'token', 'token_' + uuid.uuid4().hex[:8])
        metrics = self.form_data.get('metrics') or []
        self.metrics = []
        for metric in metrics:
            if isinstance(metric, dict):
                self.metrics.append(metric['label'])
            else:
                self.metrics.append(metric)

        self.groupby = self.form_data.get('groupby') or []
        self.time_shift = timedelta()

        self.status = None
        self.error_message = None
        self.force = force

        # Keeping track of whether some data came from cache
        # this is useful to trigerr the <CachedLabel /> when
        # in the cases where visualization have many queries
        # (FilterBox for instance)
        self._some_from_cache = False
        self._any_cache_key = None
        self._any_cached_dttm = None
        self._extra_chart_data = None
        self.datasource_data_verbose_map = None

    @staticmethod
    def handle_js_int_overflow(data):
        for d in data.get('records', dict()):
            for k, v in list(d.items()):
                if isinstance(v, int):
                    # if an int is too big for Java Script to handle
                    # convert it to a string
                    if abs(v) > JS_MAX_INTEGER:
                        d[k] = str(v)
        return data

    def run_extra_queries(self):
        """Lyfecycle method to use when more than one query is needed

        In rare-ish cases, a visualization may need to execute multiple
        queries. That is the case for FilterBox or for time comparison
        in Line chart for instance.

        In those cases, we need to make sure these queries run before the
        main `get_payload` method gets called, so that the overall caching
        metadata can be right. The way it works here is that if any of
        the previous `get_df_payload` calls hit the cache, the main
        payload's metadata will reflect that.

        The multi-query support may need more work to become a first class
        use case in the framework, and for the UI to reflect the subtleties
        (show that only some of the queries were served from cache for
        instance). In the meantime, since multi-query is rare, we treat
        it with a bit of a hack. Note that the hack became necessary
        when moving from caching the visualization's data itself, to caching
        the underlying query(ies).
        """
        pass

    def get_fillna_for_col(self, col):
        """Returns the value for use as filler for a specific Column.type"""
        if col:
            if col.is_string:
                return ' NULL'
            if col.is_time:
                return 'null'
        return self.default_fillna

    def get_fillna_for_columns(self, columns=None, session=None):
        """Returns a dict or scalar that can be passed to DataFrame.fillna"""
        if columns is None:
            return self.default_fillna
        columns_dict = {col.column_name: col for col in self.datasource.get_columns_filter(session=session).all()}
        fillna = {
            c: self.get_fillna_for_col(columns_dict.get(c))
            for c in columns
        }
        return fillna

    def format_df(self, df, query_obj, session, timestamp_format):
        if DTTM_ALIAS in df.columns:
            if timestamp_format in ('epoch_s', 'epoch_ms'):
                df[DTTM_ALIAS] = pd.to_datetime(
                    df[DTTM_ALIAS], utc=False, unit=timestamp_format[6:])
            else:
                df[DTTM_ALIAS] = pd.to_datetime(
                    df[DTTM_ALIAS], utc=False, format=timestamp_format)
            if self.datasource.offset:
                df[DTTM_ALIAS] += timedelta(hours=self.datasource.offset)
            df[DTTM_ALIAS] += self.time_shift
        self.df_metrics_to_num(df, query_obj.get('metrics') or [])

        df.replace([np.inf, -np.inf], np.nan)
        fillna = self.get_fillna_for_columns(df.columns, session=session)
        df = df.fillna(fillna)

        if self.datasource.type == 'table':
            columns_names = query_obj.get('groupby', []) + query_obj.get('columns', [])
            for column_name in columns_names:
                col = self.datasource.get_col(column_name, session=session)
                if col.is_dttm and col.python_date_format:
                    try:
                        df[column_name] = pd.to_datetime(
                            df[column_name], utc=False,
                            format=col.python_date_format, errors='coerce'
                        )
                    except ValueError as e:
                        logging.warning(e)

        return df

    def get_timestamp_format(self, query_obj, session):
        timestamp_format = None
        if self.datasource.type == 'table':
            dttm_col = self.datasource.get_col(query_obj['granularity'], session=session)
            if dttm_col:
                timestamp_format = dttm_col.python_date_format
        return timestamp_format

    def get_df(self, query_obj=None, verbose_named_columns=False, session=None):
        """Returns a pandas dataframe based on the query object"""
        if not query_obj:
            query_obj = self.query_obj()
        if not query_obj:
            return None

        self.error_msg = ''
        self.results = None

        timestamp_format = self.get_timestamp_format(query_obj, session)

        # The datasource here can be different backend but the interface is common
        self.results = self.datasource.query(query_obj, session=session)
        self.query = self.results.query
        self.status = self.results.status
        self.total_found = self.results.total_found
        self.error_message = self.results.error_message

        df = self.results.df
        # Transform the timestamp we received from database to pandas supported
        # datetime format. If no python_date_format is specified, the pattern will
        # be considered as the default ISO date format
        # If the datetime format is unix, the parse will use the corresponding
        # parsing logic.
        if df is None or df.empty:
            return pd.DataFrame()
        else:
            df = self.format_df(df, query_obj, session, timestamp_format)

        # Устанавливает название колонки - по verbose_name (по умолчанию column_name)
        if verbose_named_columns:
            df.rename(columns=self.get_datasource_data_verbose_map(session), inplace=True)

        return df

    @staticmethod
    def df_metrics_to_num(df, metrics):
        """Converting metrics to numeric when pandas.read_sql cannot"""
        for col, dtype in df.dtypes.items():
            if dtype.type == np.object_ and col in metrics:
                df[col] = pd.to_numeric(df[col])

    def convert_to_verbose_names(self, columns_or_metrics, session=None):
        """Конвертирует названия столбцов и метрик в verbose_name в соответсвии с df"""
        if isinstance(columns_or_metrics, str):
            return self.get_datasource_data_verbose_map(session).get(columns_or_metrics, columns_or_metrics)
        if columns_or_metrics:
            ds_mapping = self.get_datasource_data_verbose_map(session)
            for index, df_column in enumerate(columns_or_metrics[:]):
                column_verbose_name = ds_mapping.get(df_column)
                if column_verbose_name:
                    columns_or_metrics = [x if x != df_column else column_verbose_name for x in columns_or_metrics]

        return columns_or_metrics

    def get_time_limits_for_filter(self, query_obj):
        self.form_data['from_dttm'] = self.from_dttm.strftime('%Y-%m-%d %H:%M:%S') if self.from_dttm else ''
        self.form_data['to_dttm'] = self.to_dttm.strftime('%Y-%m-%d %H:%M:%S') if self.to_dttm else ''

    def query_obj(self):
        """Building a query object"""
        form_data = self.form_data
        gb = form_data.get('groupby') or []
        metrics = form_data.get('metrics') or []
        columns = form_data.get('columns') or []
        groupby = []
        for o in gb + columns:
            if o not in groupby:
                groupby.append(o)

        is_timeseries = self.is_timeseries
        if DTTM_ALIAS in groupby:
            groupby.remove(DTTM_ALIAS)
            is_timeseries = True

        # Add extra filters into the query form data
        merge_extra_filters(form_data)

        granularity = (
                form_data.get('granularity') or
                form_data.get('granularity_sqla')
        )
        limit = int(form_data.get('limit') or 0)
        timeseries_limit_metric = form_data.get('timeseries_limit_metric')
        row_limit = int(config.get('ROW_LIMIT') if form_data.get('row_limit') is None else form_data.get('row_limit'))
        page_length = int(form_data.get("page_length") or 0)
        page_offset = int(form_data.get("page_offset") or 0)

        # default order direction
        order_desc = form_data.get('order_desc', True)

        since = form_data.get('since', '')
        until = form_data.get('until', 'now')
        time_shift = form_data.get('time_shift', '')

        # Backward compatibility hack
        if since:
            since_words = since.split(' ')
            grains = ['days', 'years', 'hours', 'day', 'year', 'weeks']
            if (len(since_words) == 2 and since_words[1] in grains):
                since += ' ago'

        self.time_shift = utils.parse_human_timedelta(time_shift)

        since = utils.parse_human_datetime(since)
        until = utils.parse_human_datetime(until)
        from_dttm = None if since is None else (since - self.time_shift)
        to_dttm = None if until is None else (until - self.time_shift)
        if from_dttm and to_dttm and from_dttm > to_dttm:
            raise Exception(_('From date cannot be larger than to date'))

        self.from_dttm = from_dttm
        self.to_dttm = to_dttm

        form_data['from_dttm'] = from_dttm
        form_data['to_dttm'] = to_dttm

        # extras are used to query elements specific to a datasource type
        # for instance the extra where clause that applies only to Tables

        form_data['where'] = merge_where(form_data.get('where', ''), form_data.get('extra_where', ''))

        extras = {
            'where': form_data.get('where', ''),
            'having': form_data.get('having', ''),
            'having_druid': form_data.get('having_filters', []),
            'time_grain_sqla': form_data.get('time_grain_sqla', ''),
            'druid_time_origin': form_data.get('druid_time_origin', ''),
        }
        filters = form_data.get('filters', [])
        d = {
            'granularity': granularity,
            'from_dttm': from_dttm,
            'to_dttm': to_dttm,
            'is_timeseries': is_timeseries,
            'groupby': groupby,
            'metrics': metrics,
            'row_limit': row_limit,
            'page_length': page_length,
            'page_offset': page_offset,
            'filter': filters,
            'timeseries_limit': limit,
            'extras': extras,
            'timeseries_limit_metric': timeseries_limit_metric,
            'order_desc': order_desc,
            'prequeries': [],
            'is_prequery': False,
        }

        self.get_time_limits_for_filter(d)
        order_by_metric = form_data.get('order_by_metric')
        if order_by_metric:
            order_by = [(col, dir_ == 'ASC') for col, dir_ in order_by_metric]
            d['orderby'] = order_by
        elif form_data['viz_type'] == 'pivot_table':
            d['orderby'] = [(c, True) for c in form_data['groupby'] + form_data['columns']]
        return d

    @property
    def cache_timeout(self):
        if self.form_data.get('cache_timeout'):
            return int(self.form_data.get('cache_timeout'))
        if self.datasource.cache_timeout:
            return self.datasource.cache_timeout
        if (
                hasattr(self.datasource, 'database') and
                self.datasource.database.cache_timeout):
            return self.datasource.database.cache_timeout
        return config.get('CACHE_DEFAULT_TIMEOUT')

    def get_json(self):
        return json.dumps(
            self.get_payload(),
            default=utils.json_int_dttm_ser, ignore_nan=True)

    def get_hierarchy(self):
        data = list()

        for hier in self.datasource.hierarchies:
            hier_data = dict(name=hier.hier_name, id=hier.id)
            columns = []

            for c in hier.columns:
                column_data = {
                    'name': c.column.column_name,
                    'verbose_name': c.column.verbose_name,
                    'order': c.column_order,
                    'id': c.column_id,
                    'groupby': c.column.groupby,
                }
                columns.append(column_data)

            hier_data['columns'] = columns
            data.append(hier_data)

        return data

    def cache_key(self, query_obj):
        """
        The cache key is made out of the key/values in `query_obj`

        We remove datetime bounds that are hard values,
        and replace them with the use-provided inputs to bounds, which
        may we time-relative (as in "5 days ago" or "now").
        """
        cache_dict = copy.deepcopy(query_obj)

        cache_dict['user_id'] = getattr(getattr(g, 'user', None), 'id', None)

        for k in ['from_dttm', 'to_dttm']:
            del cache_dict[k]

        for k in ['since', 'until']:
            cache_dict[k] = self.form_data.get(k)

        cache_dict['datasource'] = self.datasource.uid
        json_data = self.json_dumps(cache_dict, sort_keys=True)
        return hashlib.md5(json_data.encode('utf-8')).hexdigest()

    def get_payload(self, query_obj=None, session=None):
        """Returns a payload of metadata and data"""
        self.run_extra_queries()
        payload = self.get_df_payload(query_obj, session=session)
        df = payload.get('df')
        if self.status != utils.QueryStatus.FAILED:
            if df is not None and df.empty:
                payload['error'] = str(_('No data'))
            else:
                payload['data'] = self.get_data(df, session=session)
        if 'df' in payload:
            del payload['df']
        return payload

    def get_payload_df(self, df, session=None):
        payload = self.get_df_payload_from_df(df, session=session)

        df = payload.get('df')
        if self.status != utils.QueryStatus.FAILED:
            if df is not None and df.empty:
                payload['error'] = str(_('No data'))
            else:
                payload['data'] = self.get_data(df, session=session, with_polygons=False)
        if 'df' in payload:
            del payload['df']
        return payload

    def get_df_payload_from_df(self, df, session=None):
        try:
            df.fillna(inplace=True, value='null')
        except AttributeError:
            pass

        return {
            'df': df,
        }

    def get_df_payload(self, query_obj=None, session=None):
        """Handles caching around the df payload retrieval"""
        if not query_obj:
            query_obj = self.query_obj()
        cache_key = self.cache_key(query_obj) if query_obj else None
        logging.info('Cache key: {}'.format(cache_key))
        is_loaded = False
        stacktrace = None
        df = None
        cached_dttm = datetime.utcnow().isoformat().split('.')[0]
        if cache_key and cache and not self.force:
            cache_value = cache.get(cache_key)
            if cache_value:
                stats_logger.incr('loaded_from_cache')
                try:
                    cache_value = pkl.loads(cache_value)
                    df = cache_value['df']
                    self.query = cache_value['query']
                    self.total_found = cache_value['total_found']
                    self._any_cached_dttm = cache_value['dttm']
                    self._any_cache_key = cache_key
                    self.status = utils.QueryStatus.SUCCESS
                    is_loaded = True
                except Exception as e:
                    logging.exception(e)
                    logging.error('Error reading cache: ' +
                                  utils.error_msg_from_exception(e))
                logging.info('Serving from cache')

        if query_obj and not is_loaded:
            try:
                df = self.get_df(query_obj, session=session)
                if self.status != utils.QueryStatus.FAILED:
                    stats_logger.incr('loaded_from_source')
                    is_loaded = True
            except Exception as e:
                logging.exception(e)
                if not self.error_message:
                    self.error_message = escape('{}'.format(e))
                self.status = utils.QueryStatus.FAILED
                stacktrace = traceback.format_exc()

            if (
                    is_loaded and
                    cache_key and
                    cache and
                    self.status != utils.QueryStatus.FAILED):
                try:
                    cache_value = dict(
                        dttm=cached_dttm,
                        df=df if df is not None else None,
                        query=self.query,
                        total_found=self.total_found,
                    )
                    cache_value = pkl.dumps(
                        cache_value, protocol=pkl.HIGHEST_PROTOCOL)

                    logging.info('Caching {} chars at key {}'.format(
                        len(cache_value), cache_key))

                    stats_logger.incr('set_cache_key')
                    cache.set(
                        cache_key,
                        cache_value,
                        timeout=self.cache_timeout)
                except Exception as e:
                    # cache.set call can fail if the backend is down or if
                    # the key is too large or whatever other reasons
                    logging.warning('Could not cache key {}'.format(cache_key))
                    logging.exception(e)
                    cache.delete(cache_key)
        try:
            df.fillna(inplace=True, value='null')
        except AttributeError:
            pass
        return {
            'cache_key': self._any_cache_key,
            'cached_dttm': self._any_cached_dttm,
            'cache_timeout': self.cache_timeout,
            'df': df,
            'total_found': self.total_found,
            # add table hierarchy
            'hierarchy': self.get_hierarchy(),
            'error': _(self.error_message).encode('utf-8') if self.error_message else None,
            'form_data': self.form_data,
            'is_cached': self._any_cache_key is not None,
            'query': self.query,
            'status': self.status,
            'stacktrace': stacktrace,
            'rowcount': len(df.index) if df is not None else 0,
        }

    def json_dumps(self, obj, sort_keys=False):
        return json.dumps(
            obj,
            default=utils.json_int_dttm_ser,
            ignore_nan=True,
            sort_keys=sort_keys,
        )

    @property
    def data(self):
        """This is the data object serialized to the js layer"""
        content = {
            'form_data': self.form_data,
            'token': self.token,
            'viz_name': self.viz_type,
            'filter_select_enabled': self.datasource.filter_select_enabled,
        }
        return content

    def get_csv(self):
        # remove pagination limit
        self.form_data.pop('page_length', None)
        self.form_data.pop('page_limit', None)
        self.form_data.pop('page_offset', None)
        # self.form_data['row_limit'] = None
        df = self.get_df_for_csv()
        include_index = not isinstance(df.index, pd.RangeIndex)

        conf = config.get('CSV_EXPORT')
        conf['encoding'] = 'utf-8'
        return codecs.BOM_UTF8 + df.to_csv(index=include_index, **conf).encode(conf['encoding'])

    def get_df_for_csv(self):
        return self.get_df(verbose_named_columns=True)

    def get_df_for_excel(self):
        return self.get_df(verbose_named_columns=True)

    @staticmethod
    def _add_cell_borders(writer):
        side = Side(border_style=BORDER_THIN)
        no_border = Border(left=side, right=side, top=side, bottom=side)
        # for idx, column_cells in enumerate(writer.sheets['Sheet1'].columns, 1):
        #     length = max(len(str(cell.value)) for cell in column_cells)
        #     if length > 50:
        #         length = 50
        #     writer.sheets['Sheet1'].column_dimensions[get_column_letter(idx)].width = length
        for sheet in writer.book.worksheets:
            for idx, column in enumerate(sheet.columns, 1):
                max_column = 0
                for cell in column:
                    cell.border = no_border
                    cell.alignment = Alignment(wrap_text=True, horizontal='left')
                    val_len = len(str(cell.value).strip()) + 1
                    if val_len > max_column:
                        max_column = val_len
                if max_column > 50:
                    max_column = 50
                sheet.column_dimensions[get_column_letter(idx)].width = max_column

    def get_excel(self):
        # remove pagination limit
        self.form_data.pop('page_length', None)
        self.form_data.pop('page_limit', None)
        page_offset = self.form_data.pop('page_offset', None)
        row_limit = self.form_data['row_limit']

        if row_limit and page_offset:
            self.form_data['row_limit'] = row_limit + page_offset

        df = self.get_df_for_excel()
        include_index = not isinstance(df.index, pd.RangeIndex)

        output = BytesIO()
        writer = pd.ExcelWriter(output)
        df.to_excel(writer, index=include_index, encoding="utf-8", columns=df.columns)
        self._add_cell_borders(writer)
        writer.save()
        return output.getvalue()

    def get_data(self, df, session=None):
        return []

    @property
    def json_data(self):
        return json.dumps(self.data)

    def get_datasource_data_verbose_map(self, session):
        """
        For performance reason we should retrieve datasource data once
        """
        if not self.datasource_data_verbose_map:
            self.datasource_data_verbose_map = self.datasource.data(session=session)['verbose_map']
        return self.datasource_data_verbose_map


class TableViz(BaseViz):
    """A basic html table that is sortable and searchable"""

    viz_type = 'table'
    verbose_name = _('Table View')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = False

    def should_be_timeseries(self):
        fd = self.form_data
        # TODO handle datasource-type-specific code in datasource
        conditions_met = (
                (fd.get('granularity') and fd.get('granularity') != 'all') or
                (fd.get('granularity_sqla') and fd.get('time_grain_sqla'))
        )
        # if fd.get('include_time') and not conditions_met:
        #     raise Exception(_(
        #         'Pick a granularity in the Time section or '
        #         "uncheck 'Include Time'"))
        return fd.get('include_time')

    def query_obj(self):
        d = super(TableViz, self).query_obj()
        fd = self.form_data

        if fd.get('all_columns') and (fd.get('groupby') or fd.get('metrics')):
            raise Exception(_(
                'Choose either fields to [Group By] and [Metrics] or '
                '[Columns], not both'))

        sort_by = fd.get('timeseries_limit_metric')
        if fd.get('all_columns'):
            d['columns'] = fd.get('all_columns')
            d['groupby'] = []
            order_by_cols = fd.get('order_by_cols') or []
            d['orderby'] = [json.loads(t) for t in order_by_cols]
        elif sort_by:
            if sort_by not in d['metrics']:
                d['metrics'] += [sort_by]
            d['orderby'] = [(sort_by, not fd.get('order_desc', True))]

        # Add all percent metrics that are not already in the list
        if 'percent_metrics' in fd:
            d['metrics'] = d['metrics'] + list(filter(
                lambda m: m not in d['metrics'],
                fd['percent_metrics'],
            ))

        d['is_timeseries'] = self.should_be_timeseries()
        return d

    def get_data(self, df, session=None):
        fd = self.form_data
        if (
                not self.should_be_timeseries() and
                df is not None and
                DTTM_ALIAS in df
        ):
            del df[DTTM_ALIAS]

        # Sum up and compute percentages for all percent metrics
        percent_metrics = fd.get('percent_metrics', [])
        if len(percent_metrics):
            percent_metrics = list(filter(lambda m: m in df, percent_metrics))
            metric_sums = {
                m: reduce(lambda a, b: a + b, df[m])
                for m in percent_metrics
            }
            metric_percents = {
                m: list(map(lambda a: a / metric_sums[m], df[m]))
                for m in percent_metrics
            }
            for m in percent_metrics:
                m_name = '%' + m
                df[m_name] = pd.Series(metric_percents[m], name=m_name)
            # Remove metrics that are not in the main metrics list
            for m in filter(
                    lambda m: m not in fd['metrics'] and m in df.columns,
                    percent_metrics,
            ):
                del df[m]

        data = self.handle_js_int_overflow(
            dict(
                records=df.to_dict(orient='records'),
                columns=list(df.columns),
            ))

        return data

    def json_dumps(self, obj, sort_keys=False):
        # форматированием даты в таблицах занимается фронт, выдаем только в формате timestamp
        return super(TableViz, self).json_dumps(obj)


class TimeTableViz(BaseViz):
    """A data table with rich time-series related columns"""

    viz_type = 'time_table'
    verbose_name = _('Time Table View')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = True

    def query_obj(self):
        d = super(TimeTableViz, self).query_obj()
        fd = self.form_data

        if not fd.get('metrics'):
            raise Exception(_('Pick at least one metric'))

        if fd.get('groupby') and len(fd.get('metrics')) > 1:
            raise Exception(_(
                "When using 'Group By' you are limited to use a single metric"))
        return d

    def get_data(self, df, session=None):
        fd = self.form_data
        values = self.metrics
        columns = None
        if fd.get('groupby'):
            values = self.metrics[0]
            columns = fd.get('groupby')
        pt = df.pivot_table(
            index=self.convert_to_verbose_names(DTTM_ALIAS, session=session),
            columns=self.convert_to_verbose_names(columns, session=session),
            values=self.convert_to_verbose_names(values, session=session),
            margins_name=__('All'))
        pt.index = pt.index.map(str)
        pt = pt.sort_index()
        return dict(
            records=pt.to_dict(orient='index'),
            columns=list(pt.columns),
            is_group_by=len(fd.get('groupby')) > 0,
        )


class SubPivotTableGenerator:
    '''
    Класс для создания "Сводной таблицы" во внутреннем df
    '''

    def __init__(self, indexes, values, group_field,
                 subtotal_fields, agg_func, rows, columns):
        self.indexes = indexes
        self.values = values
        self.agg_func = agg_func
        self.group_field = group_field
        self.subtotal_fields = subtotal_fields
        self.margins_name = f'‹{__("Subtotal")}›'
        self.rows = rows
        self.columns = columns

    def concat_df(self, base_df, df_to_concat):
        row_index = self.rows.index(self.group_field)
        total_col_names = ['' for _ in range(len(self.rows[row_index:]) - 1)]

        df_to_concat = df_to_concat.reindex(columns=base_df.columns)

        if len(total_col_names) > 1:
            total_col_names[0] = self.margins_name
            df_to_concat.index = [tuple(total_col_names)]
            base_df = base_df.append(df_to_concat)
        else:
            df_to_concat.index = [self.margins_name]
            base_df = base_df.append(df_to_concat)

        return base_df

    def __call__(self, sub_df, *args, **kwargs):
        '''
        :param sub_df: Под dataframe
        :return: dataframe

        Если была группировка по текущему `group_field`,
        создаем "Сводную таблицу" в под dataframe с итогами,
        удаляем строки с пустыми именами
        '''
        df = sub_df
        if self.group_field in self.subtotal_fields and self.group_field != self.rows[-1]:
            df = sub_df.pivot_table(
                index=self.indexes,
                values=self.values,
                aggfunc=self.agg_func,
                margins=not not self.columns,
                margins_name=self.margins_name,
            )
            for index, row in df.iterrows():
                if isinstance(row.name, tuple) and (
                        not any(row.name) or self.margins_name in (
                        row.name if isinstance(row.name, (tuple, list)) else [row.name])):
                    df = df.drop(index)

            if self.group_field not in self.columns:
                names_before = df.index.names

                for field in self.columns:
                    df = df.unstack(field)

                if isinstance(self.agg_func, str):
                    name_to_agg_func = {
                        'sum': df.sum,
                        'min': df.min,
                        'max': df.max,
                        'mean': df.mean,
                        'median': df.median,
                        'stdev': df.std,
                        'var': df.var
                    }
                    agg_func = name_to_agg_func.get(self.agg_func)
                    aggregated = agg_func(axis=0, numeric_only=True)
                    try:
                        agg_df = pd.DataFrame(data=aggregated).T
                    except Exception as e:
                        logging.error(e)
                        logging.error(f'data={aggregated},type={type(aggregated)}')
                        raise
                    df = self.concat_df(df.dropna(axis=1, how='all'), agg_df)

                elif isinstance(self.agg_func, dict):
                    columns_agg_df = []

                    if not df.empty:
                        for metric_name, agg_dunc in self.agg_func.items():
                            name_to_agg_func = {
                                'sum': df[[metric_name]].sum,
                                'min': df[[metric_name]].min,
                                'max': df[[metric_name]].max,
                                'mean': df[[metric_name]].mean,
                                'median': df[[metric_name]].median,
                                'stdev': df[[metric_name]].std,
                                'var': df[[metric_name]].var
                            }
                            agg_func = name_to_agg_func.get(agg_dunc)
                            aggregated = agg_func(axis=0, numeric_only=True)
                            try:
                                agg_df = pd.DataFrame(data=aggregated).T
                            except Exception as e:
                                logging.error(e)
                                logging.error(f'data={aggregated},type={type(aggregated)}')
                                raise

                            new_columns = [columns for columns in df.columns.tolist() if metric_name in columns]
                            agg_df.columns = new_columns
                            columns_agg_df.append(agg_df)

                        agg_df = pd.concat(columns_agg_df, axis=1)
                        df = self.concat_df(df.dropna(axis=1, how='all'), agg_df)

                for field in self.columns:
                    df = df.stack(field)

                df.index.names = names_before

        return df


class PivotTableViz(BaseViz):
    """A pivot table view, define your rows, columns and metrics"""

    viz_type = 'pivot_table'
    verbose_name = _('Pivot Table')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = False

    def query_obj(self):
        d = super(PivotTableViz, self).query_obj()
        groupby = self.form_data.get('groupby')
        columns = self.form_data.get('columns')
        metrics = self.form_data.get('metrics')
        if not columns:
            columns = []
        if not groupby:
            groupby = []
        if not groupby:
            raise Exception(_("Please choose at least one 'Group by' field "))
        if not metrics:
            raise Exception(_('Please choose at least one metric'))
        if (
                any(v in groupby for v in columns) or
                any(v in columns for v in groupby)):
            raise Exception(_("Group By' and 'Columns' can't overlap"))
        return d

    def _calculate_subtotals(
            self, df, indexes, columns, metrics,
            sub_totals_aggfunc, subtotals_fields, additional_fields=None):

        df = df.pivot_table(
            index=indexes,
            columns=columns,
            values=metrics,
            aggfunc=sub_totals_aggfunc,
        )
        group_by_fields = []
        all_indexes = indexes + columns

        if additional_fields is not None:
            subtotals_fields += additional_fields

        # преобразуем показатели по столбцам на показатели по строкам
        for field in columns:
            df = df.stack(field)

        # итерация по всем индексам с последующий группировкой по ним
        for group_field in all_indexes:
            group_by_fields.append(group_field)
            indexes_ = [c for c in all_indexes if c not in group_by_fields]
            if not indexes_:
                break
            df = df.groupby(group_by_fields).apply(SubPivotTableGenerator(
                indexes_,
                metrics,
                group_field,
                subtotals_fields,
                sub_totals_aggfunc,
                indexes,
                columns,
            ))

        # обратно преобразуем показатели из строк в столбцы, которые выше были преобразованы
        for field in columns:
            df = df.unstack(field)

        return df

    def get_columns(self, session=None):
        orig_columns = self.form_data.get('columns')
        return self.convert_to_verbose_names(orig_columns, session=session), orig_columns

    def get_indexes(self, session=None):
        orig_indexes = self.form_data.get('groupby')
        return self.convert_to_verbose_names(orig_indexes, session=session), orig_indexes

    def parse_totals_payload(self, payload_slice, metrics, session):

        sub_totals_metrics = {
            self.convert_to_verbose_names(metric_data['optionName'], session=session): metric_data[
                'aggregate'].lower()
            for metric_data in payload_slice
        }
        sub_totals_metrics.update(
            {metric: SUB_TOTALS_DEFAULT_AGGR_FUNC for metric in sub_totals_metrics.keys() ^ set(metrics)})
        return sub_totals_metrics

    def handle_df(self, df, session=None):
        """
        aggregations funcs: 'sum', 'min', 'max', 'mean', 'median', 'stdev', 'var'
        """
        if self.form_data.get('granularity') == 'all' and DTTM_ALIAS in df:
            del df[DTTM_ALIAS]

        df.rename(columns=self.get_datasource_data_verbose_map(session), inplace=True)
        margins_name = f'‹{__("All")}›'
        columns, _ = self.get_columns(session=session)
        indexes, _ = self.get_indexes(session=session)
        metrics = self.convert_to_verbose_names(utils.get_metric_names(self.form_data.get('metrics')), session=session)

        sub_totals_metrics_data = self.form_data.get('sub_totals_metrics')
        if not sub_totals_metrics_data:
            sub_totals_metrics = SUB_TOTALS_DEFAULT_AGGR_FUNC
        else:
            sub_totals_metrics = self.parse_totals_payload(sub_totals_metrics_data, metrics, session)
        rows_sub_totals = self.form_data.get('rows_sub_totals', True)
        if rows_sub_totals:
            sub_totals_by_rows = self.convert_to_verbose_names(self.form_data.get('sub_totals_by_rows') or [],
                                                               session=session)
            sub_totals_by_rows = list(set(indexes) & set(sub_totals_by_rows))
        else:
            sub_totals_by_rows = []

        column_sub_totals = self.form_data.get('column_sub_totals', True)
        if column_sub_totals:
            sub_totals_by_columns = self.convert_to_verbose_names(self.form_data.get('sub_totals_by_columns') or [],
                                                                  session=session)
            sub_totals_by_columns = list(set(columns) & set(sub_totals_by_columns))
        else:
            sub_totals_by_columns = []

        # если есть подытоги по столбцам
        if column_sub_totals and sub_totals_by_columns:
            df = self._calculate_subtotals(
                df=df,
                indexes=indexes,
                columns=columns,
                metrics=metrics,
                sub_totals_aggfunc=sub_totals_metrics,
                subtotals_fields=sub_totals_by_columns,
                additional_fields=sub_totals_by_rows,
            )
        # если есть подытоги по строкам
        elif rows_sub_totals and sub_totals_by_rows:
            if len(indexes) == 1:
                sub_totals_by_rows = []
            df = self._calculate_subtotals(
                df=df,
                indexes=indexes,
                metrics=metrics,
                subtotals_fields=sub_totals_by_rows,
                sub_totals_aggfunc=sub_totals_metrics,
                columns=columns,
            )
        # если не надо считать подытоги
        else:
            df = df.pivot_table(
                index=indexes,
                columns=columns,
                values=metrics,
                aggfunc=sub_totals_metrics,
            )

        df.dropna(axis=1, how='all', inplace=True)

        pivot_margins = self.form_data.get('pivot_margins', False)
        subtotal_name = f'‹{__("Subtotal")}›'

        if pivot_margins:
            # total by rows
            parsed_totals_aggr_funcs = self.parse_totals_payload(self.form_data.get('totals_agg_funcs', dict()),
                                                                 metrics, session)
            row_total_df = df.copy()

            for index, row in row_total_df.iterrows():
                if subtotal_name in (row.name if isinstance(row.name, (tuple, list)) else [row.name]):
                    row_total_df = row_total_df.drop(index)

            totals_by_rows = list()
            # calculate totals one by one (row)
            for metric, aggr_func in parsed_totals_aggr_funcs.items():
                if not columns:
                    df_by_metric = pd.DataFrame(data={metric: [getattr(row_total_df[metric], aggr_func)(axis=0)]})
                else:
                    df_by_metric = pd.DataFrame(getattr(row_total_df[metric], aggr_func)(axis=0, numeric_only=True)).T
                    # df_by_metric.rename(columns={column: (metric, column) for column in df_by_metric.columns},
                    # inplace=True)
                totals_by_rows.append(df_by_metric)
            agg_df = pd.concat(totals_by_rows, axis=1, keys=parsed_totals_aggr_funcs.keys() if columns else None)
            # agg_func = name_to_agg_func.get(pandas_aggfunc)
            # aggregated = agg_func(axis=0, numeric_only=True)
            # try:
            #     agg_df = pd.DataFrame(data=aggregated).T
            # except Exception as e:
            #     logging.error(e)
            # logging.error(f'data={aggregated},type={type(aggregated)}')
            # raise

            total_col_names = [margins_name]

            for _ in range(len(indexes) - 1):
                total_col_names.append('')

            agg_df = agg_df.reindex(columns=df.columns)

            if len(total_col_names) > 1:
                agg_df.index = [tuple(total_col_names)]
                df = df.append(agg_df)
            else:
                agg_df.index = [margins_name]
                df = df.append(agg_df)
                df.index.names = indexes

            # total by columns
            if columns:
                row_total_df = df.drop(subtotal_name, axis=1, level=len(columns))

                for metric_name, aggr_func in parsed_totals_aggr_funcs.items():
                    # agg_func = name_to_agg_func.get(pandas_aggfunc)
                    # aggregated = row_total_df[metric_name](axis=1, numeric_only=True)
                    aggregated = getattr(row_total_df[metric_name], aggr_func)(axis=1, numeric_only=True)

                    try:
                        agg_df = pd.DataFrame(data=aggregated)
                    except Exception as e:
                        logging.error(e)
                        logging.error(f'data={aggregated},type={type(aggregated)}')
                        raise

                    multiindex_columns = [
                        margins_name for _ in range(len(df[metric_name].columns.names))
                    ]
                    multiindex_columns.insert(0, metric_name)
                    df.loc[:, tuple(multiindex_columns)] = agg_df.values

        # сортируем метрики, в таком порядке, в котором выбрал пользователь
        if isinstance(df.columns, pd.MultiIndex):
            indexes_by_metric = defaultdict(list)
            for c in df.columns.tolist():
                indexes_by_metric[c[0]].append(c)
            new_order = []
            for m in metrics:
                new_order += indexes_by_metric[m]
        else:
            new_order = metrics

        df = df.reindex(new_order, axis=1)

        # Display metrics side by side with each column
        if self.form_data.get('combine_metric'):
            df = df.stack(0).unstack()

        return df

    def get_formatter(self, df_columns, session=None):
        columns, orig_columns = self.get_columns(session=session)
        indexes, orig_indexes = self.get_indexes(session=session)

        verbose_to_orig_columns = dict(zip(columns, orig_columns))
        verbose_to_orig_indexes = dict(zip(indexes, orig_indexes))

        margins_name_indexes = []

        if columns:
            columns = df_columns.get_level_values(1)
            for idx, val in enumerate(columns):
                if not isinstance(val, (tuple, list)):
                    val = [val]

                if f'‹{__("All")}›' in val:
                    margins_name_indexes.append(idx)

        formatter_class = ExtendedHTMLFormatter
        formatter_class.margins_name_indexes = margins_name_indexes
        formatter_class.verbose_to_orig_indexes = verbose_to_orig_indexes
        formatter_class.verbose_to_orig_columns = verbose_to_orig_columns

        return formatter_class

    def get_data(self, df, session=None):
        for column in df.columns:
            if is_datetime64_any_dtype(df[column]):
                df[column] = df[column].apply(
                    lambda x: x.tz_localize('utc').astimezone(tz=utils.LOCAL_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"))
        df = self.handle_df(df, session=session)

        # pd.io.formats.format.HTMLFormatter = self.get_formatter(df.columns)
        pd.io.formats.html.HTMLFormatter = self.get_formatter(df.columns, session=session)

        return dict(
            columns=list(df.columns),
            html=df.to_html(
                na_rep='',
                classes=(
                    'dataframe table table-striped table-bordered '
                    'table-condensed table-hover').split(' ')),
        )

    def get_df_for_excel(self):
        df = super(PivotTableViz, self).get_df_for_excel()
        return self.handle_df(df)

    get_df_for_csv = get_df_for_excel


class MarkupViz(BaseViz):
    """Use html or markdown to create a free form widget"""

    viz_type = 'markup'
    verbose_name = _('Markup')
    is_timeseries = False

    def query_obj(self):
        return None

    def get_df(self, query_obj=None):
        return None

    def get_data(self, df, session=None):
        markup_type = self.form_data.get('markup_type')
        code = self.form_data.get('code', '')
        if markup_type == 'markdown':
            code = markdown(code)
        return dict(html=code, theme_css=get_css_manifest_files('theme.css'))


class SeparatorViz(MarkupViz):
    """Use to create section headers in a dashboard, similar to `Markup`"""

    viz_type = 'separator'
    verbose_name = _('Separator')


class WordCloudViz(BaseViz):
    """Build a colorful word cloud

    Uses the nice library at:
    https://github.com/jasondavies/d3-cloud
    """

    viz_type = 'word_cloud'
    verbose_name = _('Word Cloud')
    is_timeseries = False

    def query_obj(self):
        d = super(WordCloudViz, self).query_obj()

        d['metrics'] = [self.form_data.get('metric')]
        d['groupby'] = [self.form_data.get('series')]
        return d

    def get_data(self, df, session=None):
        # Ordering the columns
        df = df[[self.form_data.get('series'), self.form_data.get('metric')]]
        # Labeling the columns for uniform json schema
        df.columns = ['text', 'size']
        return df.to_dict(orient='records')


class TreemapViz(BaseViz):
    """Tree map visualisation for hierarchical data."""

    viz_type = 'treemap'
    verbose_name = _('Treemap')
    credits = '<a href="https://d3js.org">d3.js</a>'
    is_timeseries = False

    def _nest(self, metric, df):
        nlevels = df.index.nlevels
        if nlevels == 1:
            result = [{'name': n, 'value': v}
                      for n, v in zip(df.index, df[metric])]
        else:
            result = [{'name': l, 'children': self._nest(metric, df.loc[l])}
                      for l in df.index.levels[0]]
        return result

    def get_data(self, df, session=None):
        df = df.set_index(self.form_data.get('groupby'))
        chart_data = [{'name': metric, 'children': self._nest(metric, df)}
                      for metric in df.columns]
        return chart_data


class CalHeatmapViz(BaseViz):
    """Calendar heatmap."""

    viz_type = 'cal_heatmap'
    verbose_name = _('Calendar Heatmap')
    credits = (
        '<a href=https://github.com/wa0x6e/cal-heatmap>cal-heatmap</a>')
    is_timeseries = True

    def get_data(self, df, session=None):
        form_data = self.form_data

        df.columns = ['timestamp', 'metric']
        timestamps = {str(obj['timestamp'].value / 10 ** 9):
                          obj.get('metric') for obj in df.to_dict('records')}

        start = utils.parse_human_datetime(form_data.get('since'))
        end = utils.parse_human_datetime(form_data.get('until'))
        domain = form_data.get('domain_granularity')
        diff_delta = rdelta.relativedelta(end, start)
        diff_secs = (end - start).total_seconds()

        if domain == 'year':
            range_ = diff_delta.years + 1
        elif domain == 'month':
            range_ = diff_delta.years * 12 + diff_delta.months + 1
        elif domain == 'week':
            range_ = diff_delta.years * 53 + diff_delta.weeks + 1
        elif domain == 'day':
            range_ = diff_secs // (24 * 60 * 60) + 1
        else:
            range_ = diff_secs // (60 * 60) + 1

        return {
            'timestamps': timestamps,
            'start': start,
            'domain': domain,
            'subdomain': form_data.get('subdomain_granularity'),
            'range': range_,
        }

    def query_obj(self):
        qry = super(CalHeatmapViz, self).query_obj()
        qry['metrics'] = [self.form_data['metric']]
        return qry


class NVD3Viz(BaseViz):
    """Base class for all nvd3 vizs"""

    credits = '<a href="http://nvd3.org/">NVD3.org</a>'
    viz_type = None
    verbose_name = 'Base NVD3 Viz'
    is_timeseries = False


class BoxPlotViz(NVD3Viz):
    """Box plot viz from ND3"""

    viz_type = 'box_plot'
    verbose_name = _('Box Plot')
    sort_series = False
    is_timeseries = True

    def to_series(self, df, classed='', title_suffix=''):
        label_sep = ' - '
        chart_data = []
        for index_value, row in zip(df.index, df.to_dict(orient='records')):
            if isinstance(index_value, tuple):
                index_value = label_sep.join(index_value)
            boxes = defaultdict(dict)
            for (label, key), value in row.items():
                if key == 'median':
                    key = 'Q2'
                boxes[label][key] = value
            for label, box in boxes.items():
                if len(self.form_data.get('metrics')) > 1:
                    # need to render data labels with metrics
                    chart_label = label_sep.join([index_value, label])
                else:
                    chart_label = index_value
                chart_data.append({
                    'label': chart_label,
                    'values': box,
                })
        return chart_data

    def get_data(self, df, session=None):
        form_data = self.form_data
        df = df.fillna(0)

        # conform to NVD3 names
        def Q1(series):  # need to be named functions - can't use lambdas
            return np.percentile(series, 25)

        def Q3(series):
            return np.percentile(series, 75)

        whisker_type = form_data.get('whisker_options')
        if whisker_type == 'Tukey':

            def whisker_high(series):
                upper_outer_lim = Q3(series) + 1.5 * (Q3(series) - Q1(series))
                series = series[series <= upper_outer_lim]
                return series[np.abs(series - upper_outer_lim).argmin()]

            def whisker_low(series):
                lower_outer_lim = Q1(series) - 1.5 * (Q3(series) - Q1(series))
                # find the closest value above the lower outer limit
                series = series[series >= lower_outer_lim]
                return series[np.abs(series - lower_outer_lim).argmin()]

        elif whisker_type == 'Min/max (no outliers)':

            def whisker_high(series):
                return series.max()

            def whisker_low(series):
                return series.min()

        elif ' percentiles' in whisker_type:
            low, high = whisker_type.replace(' percentiles', '').split('/')

            def whisker_high(series):
                return np.percentile(series, int(high))

            def whisker_low(series):
                return np.percentile(series, int(low))

        else:
            raise ValueError('Unknown whisker type: {}'.format(whisker_type))

        def outliers(series):
            above = series[series > whisker_high(series)]
            below = series[series < whisker_low(series)]
            # pandas sometimes doesn't like getting lists back here
            return set(above.tolist() + below.tolist())

        aggregate = [Q1, np.median, Q3, whisker_high, whisker_low, outliers]
        df = df.groupby(form_data.get('groupby')).agg(aggregate)
        chart_data = self.to_series(df, session=session)
        return chart_data


class BubbleViz(NVD3Viz):
    """Based on the NVD3 bubble chart"""

    viz_type = 'bubble'
    verbose_name = _('Bubble Chart')
    is_timeseries = False

    def query_obj(self):
        form_data = self.form_data
        d = super(BubbleViz, self).query_obj()
        d['groupby'] = [
            form_data.get('entity'),
        ]
        if form_data.get('series'):
            d['groupby'].append(form_data.get('series'))
        self.x_metric = form_data.get('x')
        self.y_metric = form_data.get('y')
        self.z_metric = form_data.get('size')
        self.entity = form_data.get('entity')
        self.series = form_data.get('series') or self.entity
        d['row_limit'] = form_data.get('limit')

        d['metrics'] = [
            self.z_metric,
            self.x_metric,
            self.y_metric,
        ]
        if not all(d['metrics'] + [self.entity]):
            raise Exception(_('Pick a metric for x, y and size'))
        return d

    def get_data(self, df, session=None):
        df['x'] = df[[self.x_metric]]
        df['y'] = df[[self.y_metric]]
        df['size'] = df[[self.z_metric]]
        df['shape'] = 'circle'
        df['group'] = df[[self.series]]

        series = defaultdict(list)
        for row in df.to_dict(orient='records'):
            series[row['group']].append(row)
        chart_data = []
        for k, v in series.items():
            chart_data.append({
                'key': k,
                'values': v})
        return chart_data


class BulletViz(NVD3Viz):
    """Based on the NVD3 bullet chart"""

    viz_type = 'bullet'
    verbose_name = _('Bullet Chart')
    is_timeseries = False

    def query_obj(self):
        form_data = self.form_data
        d = super(BulletViz, self).query_obj()
        self.metric = form_data.get('metric')

        def as_strings(field):
            value = form_data.get(field)
            return value.split(',') if value else []

        def as_floats(field):
            return [float(x) for x in as_strings(field)]

        self.ranges = as_floats('ranges')
        self.range_labels = as_strings('range_labels')
        self.markers = as_floats('markers')
        self.marker_labels = as_strings('marker_labels')
        self.marker_lines = as_floats('marker_lines')
        self.marker_line_labels = as_strings('marker_line_labels')

        d['metrics'] = [
            self.metric,
        ]
        if not self.metric:
            raise Exception(_('Pick a metric to display'))
        return d

    def get_data(self, df, session=None):
        df = df.fillna(0)
        df['metric'] = df[[self.metric]]
        values = df['metric'].values
        return {
            'measures': values.tolist(),
            'ranges': self.ranges or [0, values.max() * 1.1],
            'rangeLabels': self.range_labels or None,
            'markers': self.markers or None,
            'markerLabels': self.marker_labels or None,
            'markerLines': self.marker_lines or None,
            'markerLineLabels': self.marker_line_labels or None,
        }


class BigNumberViz(BaseViz):
    """Put emphasis on a single metric with this big number viz"""

    viz_type = 'big_number'
    verbose_name = _('Big Number with Trendline')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = True

    def query_obj(self):
        d = super(BigNumberViz, self).query_obj()
        metric = self.form_data.get('metric')
        if not metric:
            raise Exception(_('Pick a metric!'))
        d['metrics'] = [self.form_data.get('metric')]
        self.form_data['metric'] = metric
        return d

    def get_data(self, df, session=None):
        form_data = self.form_data
        df.sort_values(by=df.columns[0], inplace=True)
        compare_lag = form_data.get('compare_lag')
        return {
            'data': df.values.tolist(),
            'compare_lag': compare_lag,
            'compare_suffix': form_data.get('compare_suffix', ''),
        }


class BigNumberTotalViz(BaseViz):
    """Put emphasis on a single metric with this big number viz"""

    viz_type = 'big_number_total'
    verbose_name = _('Big Number')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = False

    def query_obj(self):
        d = super(BigNumberTotalViz, self).query_obj()
        metric = self.form_data.get('metric')
        if not metric:
            raise Exception(_('Pick a metric!'))
        d['metrics'] = [self.form_data.get('metric')]
        self.form_data['metric'] = metric
        return d

    def get_data(self, df, session=None):
        form_data = self.form_data
        df.sort_values(by=df.columns[0], inplace=True)
        return {
            'data': df.values.tolist(),
            'subheader': form_data.get('subheader', ''),
        }


class NVD3TimeSeriesViz(NVD3Viz):
    """A rich line chart component with tons of options"""

    viz_type = 'line'
    verbose_name = _('Time Series - Line Chart')
    sort_series = False
    is_timeseries = True

    def to_series(self, df, classed='', title_suffix=''):
        cols = []
        for col in df.columns:
            if col == '':
                cols.append('N/A')
            elif col is None:
                cols.append('NULL')
            else:
                cols.append(col)
        df.columns = cols
        series = df.to_dict('series')

        chart_data = []
        for name in df.T.index.tolist():
            ys = series[name]
            if df[name].dtype.kind not in 'biufc':
                continue
            if isinstance(name, list):
                series_title = [str(title) for title in name]
            elif isinstance(name, tuple):
                series_title = tuple(str(title) for title in name)
            else:
                series_title = str(name)
            if (
                    isinstance(series_title, (list, tuple)) and
                    len(series_title) > 1 and
                    len(self.metrics) == 1):
                # Removing metric from series name if only one metric
                series_title = series_title[1:]
            if title_suffix:
                if isinstance(series_title, string_types):
                    series_title = (series_title, title_suffix)
                elif isinstance(series_title, (list, tuple)):
                    series_title = series_title + (title_suffix,)

            values = []
            for ds in df.index:
                if ds in ys:
                    d = {
                        'x': str(ds),
                        'y': ys[ds],
                    }
                else:
                    d = {}
                values.append(d)

            d = {
                'key': series_title,
                'values': values,
            }
            if classed:
                d['classed'] = classed
            chart_data.append(d)
        return chart_data

    def process_data(self, df, aggregate=False):
        fd = self.form_data
        if fd.get('granularity') == 'all':
            raise Exception(_('Pick a time granularity for your time series'))
        if not aggregate:
            df = df.pivot_table(
                index=DTTM_ALIAS,
                columns=fd.get('groupby'),
                values=utils.get_metric_names(fd.get('metrics', [])),
                margins_name=__('All')
            )
        else:
            df = df.pivot_table(
                index=DTTM_ALIAS,
                columns=fd.get('groupby'),
                values=utils.get_metric_names(fd.get('metrics', [])),
                fill_value=0,
                aggfunc=sum,
                margins_name=__('All')
            )

        fm = fd.get('resample_fillmethod')
        if not fm:
            fm = None
        how = fd.get('resample_how')
        rule = fd.get('resample_rule')
        df.index = pd.to_datetime(df.index)
        if how and rule:
            df = df.resample(rule, how=how, fill_method=fm)
            if not fm:
                df = df.fillna(0)

        if self.sort_series:
            dfs = df.sum()
            dfs.sort_values(ascending=False, inplace=True)
            df = df[dfs.index]

        if fd.get('contribution'):
            dft = df.T
            df = (dft / dft.sum()).T

        rolling_type = fd.get('rolling_type')
        rolling_periods = int(fd.get('rolling_periods') or 0)
        min_periods = int(fd.get('min_periods') or 0)

        if rolling_type in ("mean", "std", "sum") and rolling_periods:
            kwargs = dict(window=rolling_periods, min_periods=min_periods)
            if rolling_type == "mean":
                df = df.rolling(**kwargs).mean()
            elif rolling_type == "std":
                df = df.rolling(**kwargs).std()
            elif rolling_type == "sum":
                df = df.rolling(**kwargs).sum()
        elif rolling_type == "cumsum":
            df = df.cumsum()
        if min_periods:
            df = df[min_periods:]

        num_period_compare = fd.get('num_period_compare')
        if num_period_compare:
            num_period_compare = int(num_period_compare)
            prt = fd.get('period_ratio_type')
            if prt and prt == 'growth':
                df = (df / df.shift(num_period_compare)) - 1
            elif prt and prt == 'value':
                df = df - df.shift(num_period_compare)
            else:
                df = df / df.shift(num_period_compare)

            df = df[num_period_compare:]
        return df

    def run_extra_queries(self):
        fd = self.form_data
        time_compare = fd.get('time_compare')
        if time_compare:
            query_object = self.query_obj()
            delta = utils.parse_human_timedelta(time_compare)
            query_object['inner_from_dttm'] = query_object['from_dttm']
            query_object['inner_to_dttm'] = query_object['to_dttm']

            if not query_object['from_dttm'] or not query_object['to_dttm']:
                raise Exception(_(
                    '`Since` and `Until` time bounds should be specified '
                    'when using the `Time Shift` feature.'))
            query_object['from_dttm'] -= delta
            query_object['to_dttm'] -= delta

            df2 = self.get_df_payload(query_object).get('df')
            if df2 is not None:
                df2[DTTM_ALIAS] += delta
                df2 = self.process_data(df2)
                self._extra_chart_data = self.to_series(
                    df2, classed='superset', title_suffix='---')

    def get_data(self, df, session=None):
        for column in df.columns:
            if is_datetime64_any_dtype(df[column]):
                df[column] = df[column].apply(
                    lambda x: x.tz_localize('utc').astimezone(tz=utils.LOCAL_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"))
        df = self.process_data(df)
        chart_data = self.to_series(df)

        if self._extra_chart_data:
            chart_data += self._extra_chart_data
            chart_data = sorted(chart_data, key=lambda x: tuple(x['key']))

        return chart_data


class NVD3DualLineViz(NVD3Viz):
    """A rich line chart with dual axis"""

    viz_type = 'dual_line'
    verbose_name = _('Time Series - Dual Axis Line Chart')
    sort_series = False
    is_timeseries = True

    def query_obj(self):
        d = super(NVD3DualLineViz, self).query_obj()
        m1 = self.form_data.get('metric')
        m2 = self.form_data.get('metric_2')
        d['metrics'] = [m1, m2]
        if not m1:
            raise Exception(_('Pick a metric for left axis!'))
        if not m2:
            raise Exception(_('Pick a metric for right axis!'))
        if m1 == m2:
            raise Exception(_('Please choose different metrics'
                              ' on left and right axis'))
        return d

    def to_series(self, df, classed='', session=None):
        cols = []
        for col in df.columns:
            if col == '':
                cols.append('N/A')
            elif col is None:
                cols.append('NULL')
            else:
                cols.append(col)
        df.columns = cols
        series = df.to_dict('series')
        chart_data = []
        metrics = [
            (self.form_data.get('metric')),
            self.form_data.get('metric_2'),
        ]
        metrics = self.convert_to_verbose_names(utils.get_metric_names(metrics), session=session)
        for i, m in enumerate(metrics):
            ys = series[m]
            if df[m].dtype.kind not in 'biufc':
                continue
            series_title = m
            d = {
                'key': series_title,
                'classed': classed,
                'values': [
                    {'x': ds, 'y': ys[ds] if ds in ys else None}
                    for ds in df.index
                ],
                'yAxis': i + 1,
                'type': 'line',
            }
            chart_data.append(d)
        return chart_data

    def get_data(self, df, session=None):
        fd = self.form_data
        df = df.fillna(0)

        if self.form_data.get('granularity') == 'all':
            raise Exception(_('Pick a time granularity for your time series'))

        metric = fd.get('metric')
        metric_2 = fd.get('metric_2')
        df.rename(columns=self.datasource.data(session=session)['verbose_map'], inplace=True)
        df = df.pivot_table(
            index=self.convert_to_verbose_names(DTTM_ALIAS, session=session),
            values=self.convert_to_verbose_names(utils.get_metric_names([metric, metric_2]), session=session),
            margins_name=__('All')
        )

        chart_data = self.to_series(df, session=session)
        return chart_data


class NVD3TimeSeriesBarViz(NVD3TimeSeriesViz):
    """A bar chart where the x axis is time"""

    viz_type = 'bar'
    sort_series = True
    verbose_name = _('Time Series - Bar Chart')


class NVD3TimePivotViz(NVD3TimeSeriesViz):
    """Time Series - Periodicity Pivot"""

    viz_type = 'time_pivot'
    sort_series = True
    verbose_name = _('Time Series - Period Pivot')

    def query_obj(self):
        d = super(NVD3TimePivotViz, self).query_obj()
        d['metrics'] = [self.form_data.get('metric')]
        return d

    def get_data(self, df, session=None):
        fd = self.form_data
        df = self.process_data(df)
        freq = to_offset(fd.get('freq'))
        freq.normalize = True
        df[DTTM_ALIAS] = df.index.map(freq.rollback)
        df['ranked'] = df[DTTM_ALIAS].rank(method='dense', ascending=False) - 1
        df.ranked = df.ranked.map(int)
        df['series'] = '-' + df.ranked.map(str)
        df['series'] = df['series'].str.replace('-0', 'current')
        rank_lookup = {
            row['series']: row['ranked']
            for row in df.to_dict(orient='records')
        }
        max_ts = df[DTTM_ALIAS].max()
        max_rank = df['ranked'].max()
        df[DTTM_ALIAS] = df.index + (max_ts - df[DTTM_ALIAS])
        df = df.pivot_table(
            index=DTTM_ALIAS,
            columns='series',
            values=fd.get('metric'),
            margins_name=__('All')
        )
        chart_data = self.to_series(df, session=session)
        for serie in chart_data:
            serie['rank'] = rank_lookup[serie['key']]
            serie['perc'] = 1 - (serie['rank'] / (max_rank + 1))
        return chart_data


class NVD3CompareTimeSeriesViz(NVD3TimeSeriesViz):
    """A line chart component where you can compare the % change over time"""

    viz_type = 'compare'
    verbose_name = _('Time Series - Percent Change')


class NVD3TimeSeriesStackedViz(NVD3TimeSeriesViz):
    """A rich stack area chart"""

    viz_type = 'area'
    verbose_name = _('Time Series - Stacked')
    sort_series = True


class DistributionPieViz(NVD3Viz):
    """Annoy visualization snobs with this controversial pie chart"""

    viz_type = 'pie'
    verbose_name = _('Distribution - NVD3 - Pie Chart')
    is_timeseries = False

    def get_data(self, df, session=None):
        fd = self.form_data

        columns = fd.get('columns') or []
        pt = df.pivot_table(
            index=self.groupby,
            columns=columns,
            values=self.metrics[0],
            margins_name=__('All')
        )
        pt.sort_values(by=self.metrics[0], ascending=False, inplace=True)

        chart_data = []
        for name, ys in pt.items():
            values = []
            for i, v in ys.items():
                x = i
                if not isinstance(x, (tuple, list)):
                    x = [x]

                values.append({
                    'x': [s for s in x],
                    'y': v,
                })

            chart_data = values

        return chart_data


class HistogramViz(BaseViz):
    """Histogram"""

    viz_type = 'histogram'
    verbose_name = _('Histogram')
    is_timeseries = False

    def query_obj(self):
        """Returns the query object for this visualization"""
        d = super(HistogramViz, self).query_obj()
        d['row_limit'] = self.form_data.get(
            'row_limit', int(config.get('VIZ_ROW_LIMIT')))
        numeric_columns = self.form_data.get('all_columns_x')
        if numeric_columns is None:
            raise Exception(_('Must have at least one numeric column specified'))
        self.columns = numeric_columns
        d['columns'] = numeric_columns + self.groupby
        # override groupby entry to avoid aggregation
        d['groupby'] = []
        return d

    def get_data(self, df, session=None):
        """Returns the chart data"""
        chart_data = []
        if len(self.groupby) > 0:
            groups = df.groupby(self.groupby)
        else:
            groups = [((), df)]
        for keys, data in groups:
            if isinstance(keys, str):
                keys = (keys,)
            # removing undesirable characters
            keys = [re.sub(r'\W+', r'_', k) for k in keys]
            chart_data.extend([{
                'key': '__'.join([c] + keys),
                'values': data[c].tolist()}
                for c in self.columns])
        return chart_data


class DistributionBarViz(DistributionPieViz):
    """A good old bar chart"""

    viz_type = 'dist_bar'
    verbose_name = _('Distribution - Bar Chart')
    is_timeseries = False

    def query_obj(self):
        d = super(DistributionBarViz, self).query_obj()  # noqa
        fd = self.form_data
        if (
                len(d['groupby']) <
                len(fd.get('groupby') or []) + len(fd.get('columns') or [])
        ):
            raise Exception(
                _("Can't have overlap between Series and Breakdowns"))
        if not fd.get('metrics'):
            raise Exception(_('Pick at least one metric'))
        if not fd.get('groupby'):
            raise Exception(_('Pick at least one field for [Series]'))
        return d

    def get_data(self, df, session=None):
        fd = self.form_data

        row = df.groupby(self.groupby).sum()[self.metrics[0]].copy()
        row.sort_values(ascending=False, inplace=True)
        columns = fd.get('columns') or []
        pt = df.pivot_table(
            index=self.groupby,
            columns=columns,
            values=self.metrics,
            margins_name=__('All')
        )
        if fd.get('contribution'):
            pt = pt.fillna(0)
            pt = pt.T
            pt = (pt / pt.sum()).T
        pt = pt.reindex(row.index)
        chart_data = []
        for name, ys in pt.items():
            if pt[name].dtype.kind not in 'biufc' or name in self.groupby:
                continue
            if isinstance(name, string_types):
                series_title = name
            elif len(self.metrics) > 1:
                series_title = ', '.join(name)
            else:
                l = [str(s) for s in name[1:]]  # noqa: E741
                series_title = ', '.join(l)
            values = []
            for i, v in ys.items():
                x = i
                if not isinstance(x, (tuple, list)):
                    x = [x]

                values.append({
                    'x': [s for s in x],
                    'y': v,
                })
            d = {
                'key': series_title,
                'values': values,
            }
            chart_data.append(d)
        return chart_data


class SunburstViz(BaseViz):
    """A multi level sunburst chart"""

    viz_type = 'sunburst'
    verbose_name = _('Sunburst')
    is_timeseries = False
    credits = (
        'Kerry Rodden '
        '@<a href="https://bl.ocks.org/kerryrodden/7090426">bl.ocks.org</a>')

    def get_data(self, df, session=None):
        fd = self.form_data
        cols = fd.get('groupby')
        metric = fd.get('metric')
        secondary_metric = fd.get('secondary_metric')
        if metric == secondary_metric or secondary_metric is None:
            df.columns = cols + ['m1']
            df['m2'] = df['m1']
        return json.loads(df.to_json(orient='values'))

    def query_obj(self):
        qry = super(SunburstViz, self).query_obj()
        fd = self.form_data
        qry['metrics'] = [fd['metric']]
        secondary_metric = fd.get('secondary_metric')
        if secondary_metric and secondary_metric != fd['metric']:
            qry['metrics'].append(secondary_metric)
        return qry


class SankeyViz(BaseViz):
    """A Sankey diagram that requires a parent-child dataset"""

    viz_type = 'sankey'
    verbose_name = _('Sankey')
    is_timeseries = False
    credits = '<a href="https://www.npmjs.com/package/d3-sankey">d3-sankey on npm</a>'

    def query_obj(self):
        qry = super(SankeyViz, self).query_obj()
        if len(qry['groupby']) != 2:
            raise Exception(_('Pick exactly 2 columns as [Source / Target]'))
        qry['metrics'] = [
            self.form_data['metric']]
        return qry

    def get_data(self, df, session=None):
        df.columns = ['source', 'target', 'value']
        recs = df.to_dict(orient='records')

        hierarchy = defaultdict(set)
        for row in recs:
            hierarchy[row['source']].add(row['target'])

        def find_cycle(g):
            """Whether there's a cycle in a directed graph"""
            path = set()

            def visit(vertex):
                path.add(vertex)
                for neighbour in g.get(vertex, ()):
                    if neighbour in path or visit(neighbour):
                        return (vertex, neighbour)
                path.remove(vertex)

            for v in g:
                cycle = visit(v)
                if cycle:
                    return cycle

        cycle = find_cycle(hierarchy)
        if cycle:
            raise Exception(_(
                "There's a loop in your Sankey, please provide a tree. "
                "Here's a faulty link: {}").format(cycle))
        return recs


class DirectedForceViz(BaseViz):
    """An animated directed force layout graph visualization"""

    viz_type = 'directed_force'
    verbose_name = _('Directed Force Layout')
    credits = 'd3noob @<a href="http://bl.ocks.org/d3noob/5141278">bl.ocks.org</a>'
    is_timeseries = False

    def query_obj(self):
        qry = super(DirectedForceViz, self).query_obj()
        if len(self.form_data['groupby']) != 2:
            raise Exception(_("Pick exactly 2 columns to 'Group By'"))
        qry['metrics'] = [self.form_data['metric']]
        return qry

    def get_data(self, df, session=None):
        df.columns = ['source', 'target', 'value']
        return df.to_dict(orient='records')


class ChordViz(BaseViz):
    """A Chord diagram"""

    viz_type = 'chord'
    verbose_name = _('Directed Force Layout')
    credits = '<a href="https://github.com/d3/d3-chord">Bostock</a>'
    is_timeseries = False

    def query_obj(self):
        qry = super(ChordViz, self).query_obj()
        fd = self.form_data
        qry['groupby'] = [fd.get('groupby'), fd.get('columns')]
        qry['metrics'] = [fd.get('metric')]
        return qry

    def get_data(self, df, session=None):
        df.columns = ['source', 'target', 'value']

        # Preparing a symetrical matrix like d3.chords calls for
        nodes = list(set(df['source']) | set(df['target']))
        matrix = {}
        for source, target in product(nodes, nodes):
            matrix[(source, target)] = 0
        for source, target, value in df.to_records(index=False):
            matrix[(source, target)] = value
        m = [[matrix[(n1, n2)] for n1 in nodes] for n2 in nodes]
        return {
            'nodes': list(nodes),
            'matrix': m,
        }


class CountryMapViz(BaseViz):
    """A country centric"""

    viz_type = 'country_map'
    verbose_name = _('Country Map')
    is_timeseries = False
    credits = 'From bl.ocks.org By john-guerra'

    def query_obj(self):
        qry = super(CountryMapViz, self).query_obj()
        qry['metrics'] = [
            self.form_data['metric']]
        qry['groupby'] = [self.form_data['entity']]
        return qry

    def get_data(self, df, session=None):
        fd = self.form_data
        cols = [fd.get('entity')]
        metric = fd.get('metric')
        cols += [metric]
        ndf = df[cols]
        df = ndf
        df.columns = ['country_id', 'metric']
        d = df.to_dict(orient='records')
        return d


class WorldMapViz(BaseViz):
    """A country centric world map"""

    viz_type = 'world_map'
    verbose_name = _('World Map')
    is_timeseries = False
    credits = 'datamaps on <a href="https://www.npmjs.com/package/datamaps">npm</a>'

    def query_obj(self):
        qry = super(WorldMapViz, self).query_obj()
        qry['metrics'] = [
            self.form_data['metric'], self.form_data['secondary_metric']]
        qry['groupby'] = [self.form_data['entity']]
        return qry

    def get_data(self, df, session=None):
        from superset.data import countries
        fd = self.form_data
        cols = [fd.get('entity')]
        metric = fd.get('metric')
        secondary_metric = fd.get('secondary_metric')
        if metric == secondary_metric:
            ndf = df[cols]
            # df[metric] will be a DataFrame
            # because there are duplicate column names
            ndf['m1'] = df[metric].iloc[:, 0]
            ndf['m2'] = ndf['m1']
        else:
            cols += [metric, secondary_metric]
            ndf = df[cols]
        df = ndf
        df.columns = ['country', 'm1', 'm2']
        d = df.to_dict(orient='records')
        for row in d:
            country = None
            if isinstance(row['country'], string_types):
                country = countries.get(
                    fd.get('country_fieldtype'), row['country'])

            if country:
                row['country'] = country['cca3']
                row['latitude'] = country['lat']
                row['longitude'] = country['lng']
                row['name'] = country['name']
            else:
                row['country'] = 'XXX'
        return d


class FilterBoxViz(BaseViz):
    """A multi filter, multi-choice filter box to make dashboards interactive"""

    viz_type = 'filter_box'
    verbose_name = _('Filters')
    is_timeseries = False
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    cache_type = 'get_data'

    def query_obj(self):
        return None

    def run_extra_queries(self):
        qry = self.filter_query_obj()
        filters = [g for g in self.form_data['groupby']]
        self.dataframes = {}
        for flt in filters:
            qry['groupby'] = [flt]
            df = self.get_df_payload(query_obj=qry).get('df')
            self.dataframes[flt] = df

    def filter_query_obj(self):
        qry = super(FilterBoxViz, self).query_obj()
        groupby = self.form_data.get('groupby')
        if len(groupby) < 1 and not self.form_data.get('date_filter'):
            raise Exception(_('Pick at least one filter field'))
        qry['metrics'] = [
            self.form_data['metric']]
        return qry

    def get_data(self, df, session=None):
        d = {}

        filters = [g for g in self.form_data['groupby']]

        # for flt in filters:
        #     df = self.dataframes[flt]
        #     d[flt] = []
        #
        #     print(df)
        #     print(df.__dict__)
        #
        #     for row in df.itertuples(index=False):
        #         try:
        #             d[flt].append({'id': row[0], 'text': row[0], 'filter': flt, 'metric': row[1]})
        #         except IndexError:
        #             pass

        qry = self.filter_query_obj()
        for flt in filters:
            qry_ = qry.copy()
            qry_['groupby'] = [flt]

            df = super(FilterBoxViz, self).get_df(qry_)

            d[flt] = []

            for row in df.itertuples(index=False):
                try:
                    d[flt].append({'id': row[0], 'text': row[0], 'filter': flt, 'metric': row[1]})
                except IndexError:
                    pass

        return d


class IFrameViz(BaseViz):
    """You can squeeze just about anything in this iFrame component"""

    viz_type = 'iframe'
    verbose_name = _('iFrame')
    credits = 'a <a href="https://github.com/airbnb/superset">Superset</a> original'
    is_timeseries = False

    def query_obj(self):
        return None

    def get_df(self, query_obj=None):
        return None


class ParallelCoordinatesViz(BaseViz):
    """Interactive parallel coordinate implementation

    Uses this amazing javascript library
    https://github.com/syntagmatic/parallel-coordinates
    """

    viz_type = 'para'
    verbose_name = _('Parallel Coordinates')
    credits = (
        '<a href="https://syntagmatic.github.io/parallel-coordinates/">'
        "Syntagmatic's library</a>")
    is_timeseries = False

    def query_obj(self):
        d = super(ParallelCoordinatesViz, self).query_obj()
        fd = self.form_data
        d['metrics'] = copy.copy(fd.get('metrics'))
        second = fd.get('secondary_metric')
        if second not in d['metrics']:
            d['metrics'] += [second]
        d['groupby'] = [fd.get('series')]
        return d

    def get_data(self, df, session=None):
        return df.to_dict(orient='records')


class HeatmapViz(BaseViz):
    """A nice heatmap visualization that support high density through canvas"""

    viz_type = 'heatmap'
    verbose_name = _('Heatmap')
    is_timeseries = False
    credits = (
        'inspired from mbostock @<a href="http://bl.ocks.org/mbostock/3074470">'
        'bl.ocks.org</a>')

    def query_obj(self):
        d = super(HeatmapViz, self).query_obj()
        fd = self.form_data
        d['metrics'] = [fd.get('metric')]
        d['groupby'] = [fd.get('all_columns_x'), fd.get('all_columns_y')]
        return d

    def get_data(self, df, session=None):
        fd = self.form_data
        x = fd.get('all_columns_x')
        y = fd.get('all_columns_y')
        v = fd.get('metric')
        if x == y:
            df.columns = ['x', 'y', 'v']
        else:
            df = df[[x, y, v]]
            df.columns = ['x', 'y', 'v']
        norm = fd.get('normalize_across')
        overall = False
        max_ = df.v.max()
        min_ = df.v.min()
        bounds = fd.get('y_axis_bounds')
        if bounds and bounds[0] is not None:
            min_ = bounds[0]
        if bounds and bounds[1] is not None:
            max_ = bounds[1]
        if norm == 'heatmap':
            overall = True
        else:
            gb = df.groupby(norm, group_keys=False)
            if len(gb) <= 1:
                overall = True
            else:
                df['perc'] = (
                    gb.apply(
                        lambda x: (x.v - x.v.min()) / (x.v.max() - x.v.min()))
                )
        if overall:
            df['perc'] = (df.v - min_) / (max_ - min_)
        return {
            'records': df.to_dict(orient='records'),
            'extents': [min_, max_],
        }


class HorizonViz(NVD3TimeSeriesViz):
    """Horizon chart

    https://www.npmjs.com/package/d3-horizon-chart
    """

    viz_type = 'horizon'
    verbose_name = _('Horizon Charts')
    credits = (
        '<a href="https://www.npmjs.com/package/d3-horizon-chart">'
        'd3-horizon-chart</a>')


class MapboxViz(BaseViz):
    """Rich maps made with Mapbox"""

    viz_type = 'mapbox'
    verbose_name = _('Mapbox')
    is_timeseries = False
    credits = (
        '<a href=https://www.mapbox.com/mapbox-gl-js/api/>Mapbox GL JS</a>')

    def query_obj(self):
        d = super(MapboxViz, self).query_obj()
        fd = self.form_data
        label_col = fd.get('mapbox_label')

        if not fd.get('groupby'):
            d['columns'] = [fd.get('all_columns_x'), fd.get('all_columns_y')]

            if label_col and len(label_col) >= 1:
                if label_col[0] == 'count':
                    raise Exception(_(
                        "Must have a [Group By] column to have 'count' as the [Label]"))
                d['columns'].append(label_col[0])

            if fd.get('point_radius') != 'Auto':
                d['columns'].append(fd.get('point_radius'))

            d['columns'] = list(set(d['columns']))
        else:
            # Ensuring columns chosen are all in group by
            if (label_col and len(label_col) >= 1 and
                    label_col[0] != 'count' and
                    label_col[0] not in fd.get('groupby')):
                raise Exception(_(
                    'Choice of [Label] must be present in [Group By]'))

            if (fd.get('point_radius') != 'Auto' and
                    fd.get('point_radius') not in fd.get('groupby')):
                raise Exception(_(
                    'Choice of [Point Radius] must be present in [Group By]'))

            if (fd.get('all_columns_x') not in fd.get('groupby') or
                    fd.get('all_columns_y') not in fd.get('groupby')):
                raise Exception(_(
                    '[Longitude] and [Latitude] columns must be present in [Group By]'))
        return d

    def get_data(self, df, session=None):
        if df is None:
            return None
        fd = self.form_data
        label_col = fd.get('mapbox_label')
        custom_metric = label_col and len(label_col) >= 1
        metric_col = [None] * len(df.index)
        if custom_metric:
            if label_col[0] == fd.get('all_columns_x'):
                metric_col = df[fd.get('all_columns_x')]
            elif label_col[0] == fd.get('all_columns_y'):
                metric_col = df[fd.get('all_columns_y')]
            else:
                metric_col = df[label_col[0]]
        point_radius_col = (
            [None] * len(df.index)
            if fd.get('point_radius') == 'Auto'
            else df[fd.get('point_radius')])

        # using geoJSON formatting
        geo_json = {
            'type': 'FeatureCollection',
            'features': [
                {
                    'type': 'Feature',
                    'properties': {
                        'metric': metric,
                        'radius': point_radius,
                    },
                    'geometry': {
                        'type': 'Point',
                        'coordinates': [lon, lat],
                    },
                }
                for lon, lat, metric, point_radius
                in zip(
                    df[fd.get('all_columns_x')],
                    df[fd.get('all_columns_y')],
                    metric_col, point_radius_col)
            ],
        }

        return {
            'geoJSON': geo_json,
            'customMetric': custom_metric,
            'mapboxApiKey': config.get('MAPBOX_API_KEY'),
            'mapStyle': fd.get('mapbox_style'),
            'aggregatorName': fd.get('pandas_aggfunc'),
            'clusteringRadius': fd.get('clustering_radius'),
            'pointRadiusUnit': fd.get('point_radius_unit'),
            'globalOpacity': fd.get('global_opacity'),
            'viewportLongitude': fd.get('viewport_longitude'),
            'viewportLatitude': fd.get('viewport_latitude'),
            'viewportZoom': fd.get('viewport_zoom'),
            'renderWhileDragging': fd.get('render_while_dragging'),
            'tooltip': fd.get('rich_tooltip'),
            'color': fd.get('mapbox_color'),
        }


class DeckGLMultiLayer(BaseViz):
    """Pile on multiple DeckGL layers"""

    viz_type = 'deck_multi'
    verbose_name = _('Deck.gl - Multiple Layers')

    is_timeseries = False
    credits = '<a href="https://uber.github.io/deck.gl/">deck.gl</a>'

    def query_obj(self):
        return None

    def get_data(self, df, session=None):
        fd = self.form_data
        # Late imports to avoid circular import issues
        from superset.models.core import Slice
        from superset import db
        slice_ids = fd.get('deck_slices')
        slices = db.session.query(Slice).filter(Slice.id.in_(slice_ids)).all()
        return {
            'mapboxApiKey': config.get('MAPBOX_API_KEY'),
            'slices': [slc.data for slc in slices],
        }


class BaseDeckGLViz(BaseViz):
    """Base class for deck.gl visualizations"""

    is_timeseries = False
    credits = '<a href="https://uber.github.io/deck.gl/">deck.gl</a>'
    spatial_control_keys = []

    def get_metrics(self):
        self.metric = self.form_data.get('size')
        return [self.metric] if self.metric else []

    def process_spatial_query_obj(self, key, group_by):
        spatial = self.form_data.get(key)
        if spatial is None:
            raise ValueError(_('Bad spatial key'))

        if spatial.get('type') == 'latlong':
            group_by += [spatial.get('lonCol')]
            group_by += [spatial.get('latCol')]
        elif spatial.get('type') == 'delimited':
            group_by += [spatial.get('lonlatCol')]
        elif spatial.get('type') == 'geohash':
            group_by += [spatial.get('geohashCol')]

    def process_spatial_data_obj(self, key, df):
        spatial = self.form_data.get(key)
        if spatial is None:
            raise ValueError(_('Bad spatial key'))
        if spatial.get('type') == 'latlong':
            df[key] = list(zip(
                pd.to_numeric(df[spatial.get('lonCol')], errors='coerce'),
                pd.to_numeric(df[spatial.get('latCol')], errors='coerce'),
            ))
        elif spatial.get('type') == 'delimited':

            def tupleify(s):
                p = Point(s)
                return (p.latitude, p.longitude)

            df[key] = df[spatial.get('lonlatCol')].apply(tupleify)

            if spatial.get('reverseCheckbox'):
                df[key] = [
                    tuple(reversed(o)) if isinstance(o, (list, tuple)) else (0, 0)
                    for o in df[key]
                ]
            del df[spatial.get('lonlatCol')]
        elif spatial.get('type') == 'geohash':
            latlong = df[spatial.get('geohashCol')].map(geohash.decode)
            df[key] = list(zip(latlong.apply(lambda x: x[0]),
                               latlong.apply(lambda x: x[1])))
            del df[spatial.get('geohashCol')]
        return df

    def query_obj(self):
        d = super(BaseDeckGLViz, self).query_obj()
        fd = self.form_data
        gb = []

        for key in self.spatial_control_keys:
            self.process_spatial_query_obj(key, gb)

        if fd.get('dimension'):
            gb += [fd.get('dimension')]

        if fd.get('js_columns'):
            gb += fd.get('js_columns')
        metrics = self.get_metrics()
        if metrics:
            d['groupby'] = gb
            d['metrics'] = self.get_metrics()
        else:
            d['columns'] = gb

        return d

    def get_js_columns(self, d):
        cols = self.form_data.get('js_columns') or []
        return {col: d.get(col) for col in cols}

    def get_data(self, df, session=None):
        if df is None:
            return None
        for key in self.spatial_control_keys:
            df = self.process_spatial_data_obj(key, df)

        features = []
        for d in df.to_dict(orient='records'):
            feature = self.get_properties(d)
            extra_props = self.get_js_columns(d)
            if extra_props:
                feature['extraProps'] = extra_props
            features.append(feature)

        return {
            'features': features,
            'mapboxApiKey': config.get('MAPBOX_API_KEY'),
        }

    def get_properties(self, d):
        raise NotImplementedError()


class DeckScatterViz(BaseDeckGLViz):
    """deck.gl's ScatterLayer"""

    viz_type = 'deck_scatter'
    verbose_name = _('Deck.gl - Scatter plot')
    spatial_control_keys = ['spatial']
    is_timeseries = True

    def query_obj(self):
        fd = self.form_data
        self.is_timeseries = fd.get('time_grain_sqla') or fd.get('granularity')
        self.point_radius_fixed = (
                fd.get('point_radius_fixed') or {'type': 'fix', 'value': 500})
        return super(DeckScatterViz, self).query_obj()

    def get_metrics(self):
        self.metric = None
        if self.point_radius_fixed.get('type') == 'metric':
            self.metric = self.point_radius_fixed.get('value')
            return [self.metric]
        return None

    def get_properties(self, d):
        return {
            'metric': d.get(self.metric),
            'radius': self.fixed_value if self.fixed_value else d.get(self.metric),
            'cat_color': d.get(self.dim) if self.dim else None,
            'position': d.get('spatial'),
            '__timestamp': d.get(DTTM_ALIAS) or d.get('__time'),
        }

    def get_data(self, df, session=None):
        fd = self.form_data
        self.point_radius_fixed = fd.get('point_radius_fixed')
        self.fixed_value = None
        self.dim = self.form_data.get('dimension')
        if self.point_radius_fixed.get('type') != 'metric':
            self.fixed_value = self.point_radius_fixed.get('value')
        return super(DeckScatterViz, self).get_data(df, session=session)


class BubbleMapVisualization(DeckScatterViz):
    """deck.gl's ScatterLayer"""

    viz_type = 'bubble_map'
    verbose_name = _('Bubble map')
    spatial_control_keys = ['spatial']
    is_timeseries = True

    def query_obj(self):
        fd = self.form_data
        self.is_timeseries = fd.get('time_grain_sqla') or fd.get('granularity')
        self.point_name = fd.get('pointName')
        self.icon_field = fd.get('icon_field')
        self.aggregation_by_area = fd.get('aggregation_by_area')
        self.polygon_id = fd.get('polygon_id')
        # self.point_value = fd.get('pointValue')
        query_obj = super(BaseDeckGLViz, self).query_obj()
        # get long/lat list
        _fields = self.process_long_lat_query_obj()

        _custom_fields = []

        self.lat_field = _fields[0]
        self.lng_field = _fields[1]

        # also select another one column
        if self.point_name:
            _fields.append(self.point_name)
        # if icon field exists append to others fields
        if self.icon_field:
            _fields.append(self.icon_field)
        else:
            icon_field_default_name = 'icon_field'
            if self.datasource.database.db_engine_spec.engine == 'clickhouse':
                _custom_fields.append(f'CAST(NULL AS Nullable(String)) AS {icon_field_default_name}')
            else:
                _custom_fields.append(f'NULL AS {icon_field_default_name}')
            self.icon_field = icon_field_default_name
        # if groupby is not empty (values in form_data)
        if query_obj['groupby']:
            # also grouping longitude and latitude
            query_obj['groupby'] = _fields + query_obj['groupby']
        # use query without grouping
        else:
            query_obj['columns'] = _fields

        query_obj['custom_columns'] = _custom_fields

        metrics = self.get_metrics()
        if metrics:
            # as metrics may be used without grouping we should check groupby and columns lists
            groupby_list = query_obj.get('groupby', list())
            for field in query_obj.get('columns', list()):
                if field not in groupby_list:
                    groupby_list.append(field)

            query_obj['groupby'] = groupby_list
            query_obj['metrics'] = metrics

        if self.aggregation_by_area and self.polygon_id:
            query_obj["text_join"] = self.get_area_join(query_obj.get('columns'))

        return query_obj

    def get_area_join(self, query_columns):
        from superset import db
        from superset.models.core import GeoPoligons
        session = db.session

        coords_separator = ','
        lat_field = '{' + self.lat_field + '}'
        lng_field = '{' + self.lng_field + '}'

        areas_subq = session.query(
            func.json_array_elements(GeoPoligons.content).op('->')('center').op('->>')(0).label('center_x'),
            func.json_array_elements(GeoPoligons.content).op('->')('center').op('->>')(1).label('center_y'),
            func.json_array_elements(GeoPoligons.content).op('->')('polygon').label('polygon')
        ).filter(GeoPoligons.id == self.polygon_id).subquery('areas')

        # getting area coordinates as several columns
        areas_with_coordinates_subq = session.query(
            func.concat(areas_subq.c.center_x, coords_separator, areas_subq.c.center_y).label('center'),
            func.json_array_elements(areas_subq.c.polygon).op('->>')(0).label('lat'),
            func.json_array_elements(areas_subq.c.polygon).op('->>')(1).label('lng')
        ).subquery('area_coordinates')

        # building polygon object for each area to check if area contains any points
        areas_with_polygons = session.query(
            areas_with_coordinates_subq.c.center.label('center'),
            func.st_makepolygon(
                func.st_makeline(
                    func.st_makepoint(
                        areas_with_coordinates_subq.c.lat.cast(Float),
                        areas_with_coordinates_subq.c.lng.cast(Float)
                    )
                )
            ).label('polygon')
        ).group_by(areas_with_coordinates_subq.c.center).subquery('area_polygons')

        polygons_with_center_coords = session.query(
            func.string_to_array(areas_with_polygons.c.center, coords_separator).label('center_coords'),
            areas_with_polygons.c.polygon
        ).subquery('polygons_with_center_coords')

        onclause = func.st_contains(
            text(polygons_with_center_coords.c.polygon.name),
            func.st_makepoint(text(lat_field), text(lng_field))
        )
        replace_lat = case(
            [(column('center_coords')==None, text(lat_field))],
            else_=column('center_coords', type_=ARRAY(item_type=String))[1].cast(Float))
        replace_lng = case(
            [(column('center_coords')==None, text(lng_field))],
            else_=column('center_coords', type_=ARRAY(item_type=String))[2].cast(Float))

        return {
            'join_with': str(polygons_with_center_coords.compile(compile_kwargs={"literal_binds": True})),
            'on': str(onclause.compile(compile_kwargs={"literal_binds": True})),
            'columns': ['center_coords', 'polygon'],
            'query_columns': query_columns,
            'replace_query_columns': {
                self.lat_field: str(replace_lat.compile(compile_kwargs={'literal_binds': True})),
                self.lng_field: str(replace_lng.compile(compile_kwargs={'literal_binds': True}))
            }
        }

    def get_metrics(self):
        return self.form_data.get('metrics', list())

    def get_properties(self, df) -> dict:
        data = {
            'metric': [dict(name=metric, value=df.get(metric)) for metric in self.metrics],
            'groupby': [dict(name=field, value=df.get(field)) for field in self.groupby],
            # 'radius': self.fixed_value if self.fixed_value else df.get(self.metric),
            # 'cat_color': df.get(self.dim) if self.dim else None,
            'position': (df.get(self.form_data.get('latitude')), df.get(self.form_data.get('longitude'))),
            # '__timestamp': df.get(DTTM_ALIAS) or df.get('__time'),
            'pointName': df.get(self.point_name),
            # 'pointValue': df.get(self.point_value)
        }
        if self.icon_field:
            data[self.icon_field] = self.get_icon_path(df.get(self.icon_field))
        return data

    def _set_polygons_data(self, data, session=None):
        from superset import db
        from superset.models.core import GeoPoligons

        if session is None:
            session = db.session
        else:
            session = session

        polygon_id = self.form_data.get('polygon_id')

        if not polygon_id:
            areas = None
        else:
            areas = (
                session.query(GeoPoligons.content)
                .filter(GeoPoligons.id == polygon_id)
                .first()
            )

        data['areas'] = areas

        return data

    @staticmethod
    @lru_cache(maxsize=16)
    def get_icon_path(icon):
        return f'{YANDEX_ICONS}{icon}' if icon else None

    def process_long_lat_query_obj(self) -> list:
        latitude = self.form_data.get('latitude')
        longitude = self.form_data.get('longitude')
        if latitude is None or longitude is None:
            raise ValueError(_('Bad longitude or latitude key'))
        return [latitude, longitude]

    def process_spatial_data_obj(self, key, df):
        latitude = self.form_data.get('latitude')
        longitude = self.form_data.get('longitude')
        if latitude is None or longitude is None:
            raise ValueError(_('Bad longitude or latitude key'))
        if latitude in df.keys() and longitude in df.keys():
            df[key] = list(zip(
                pd.to_numeric(df[longitude], errors='coerce'),
                pd.to_numeric(df[latitude], errors='coerce'),
            ))
        return df

    def get_data(self, df, session=None, with_polygons=True):
        data = super(DeckScatterViz, self).get_data(df, session=session)

        if with_polygons:
            data = self._set_polygons_data(data, session)

        return data

    def get_payload_with_parsing(self, query_obj=None, session=None):
        """Returns a payload of metadata and data"""
        self.run_extra_queries()
        payload = self.get_df_payload(query_obj, session=session)
        df = payload.get('df')
        if self.status != utils.QueryStatus.FAILED:
            if df is not None and df.empty:
                payload['error'] = str(_('No data'))
            else:
                payload['data'] = self.get_data_with_parsing(df, session=session)
        if 'df' in payload:
            del payload['df']
        return payload

    def get_data_with_parsing(self, df, session=None):
        data = self.get_data(df, session=session)
        features = data.get('features')
        query_obj_first = self.query_obj()
        base_item = {}
        if features and isinstance(features, list):
            columns = list(features[0])
            columns = columns[len(columns) - 3:]
            for item in features:
                key_name = []
                for c in columns:
                    key = item.pop(c)
                    if type(key) == list:
                        key = tuple(key)
                    if not isinstance(key, dict):
                        key_name.append(key)
                key_name = tuple(key_name)
                if key_name in base_item:
                    base_item[key_name]['items'].append(item)
                else:
                    try:
                        base_item[key_name] = {'items': [item], 'metric': dict()}
                    except:
                        pass
        if len(query_obj_first['groupby']) > 3:
            query_obj_first['groupby'] = query_obj_first['groupby'][0:3]
            timestamp_format = self.get_timestamp_format(query_obj_first, session)
            results = self.datasource.query(query_obj_first, session=session)
            df = results.df
            df = self.format_df(df, query_obj_first, session, timestamp_format)
            columns = list(df.columns)
            for k, v in base_item.items():
                cords = k[0]
                point_name = k[1]
                item = df[
                    (df[columns[0]] == cords[0]) & (df[columns[1]] == cords[1]) & (df[columns[2]] == point_name)
                    ].to_dict(orient='index')
                item_keys = list(item)
                if item_keys:
                    metrics = item.get(item_keys[0])
                    main_metric_label = query_obj_first['metrics'][0]
                    if type(main_metric_label) == dict:
                        main_metric_label = main_metric_label.get('label')
                    [metrics.pop(col) for col in columns[:4] if col != main_metric_label]
                    v['metric'] = [{'name': mk, 'value': mv} for mk, mv in metrics.items()]
            base_item_list = [
                {
                    'position': k[0],
                    'pointName': k[1],
                    'map_icon': k[2],
                    'items': [{'pointName': k[1], **item} for item in v['items']],
                    'metric': v.get('metric')
                }
                for k, v in base_item.items()
                if v
            ]
        else:
            base_item_list = [
                {
                    'position': k[0],
                    'pointName': k[1],
                    'map_icon': k[2],
                    'items': [{'pointName': k[1], **item} for item in v['items']],
                    'metric': v['items'][0].get('metric')}
                for k, v in base_item.items()
                if v
            ]
        data['features'] = base_item_list
        return data


class YandexHeatMapVisualization(BubbleMapVisualization):
    viz_type = 'yandex_heat_map'
    verbose_name = _('Yandex heat map')
    spatial_control_keys = ['spatial']
    is_timeseries = True
    max_min_avg_gen = None
    generator_started = False
    max_min_avg = dict()

    def get_properties(self, df) -> dict:
        data = super(YandexHeatMapVisualization, self).get_properties(df)
        _metrics = data['metric']
        if not self.generator_started:
            self.max_min_avg_gen = self.calculate_max_min_avg(_metrics)
            self.max_min_avg_gen.send(None)
        else:
            self.max_min_avg_gen.send(_metrics)
        return data

    def calculate_max_min_avg(self, metrics_list: list):
        """
        Generator for calculating min, max and avg value of each metric
        """
        self.generator_started = True
        count = 0
        self.max_min_avg = {metric_dict['name']: dict(max_val=metric_dict['value'], min_val=metric_dict['value'],
                                                      avg_val=metric_dict['value']) for metric_dict in
                            metrics_list}
        metric_sum = {metric['name']: 0 for metric in metrics_list}
        while True:
            for metric_dict in metrics_list:
                metric, val = metric_dict['name'], metric_dict['value']
                _metric = self.max_min_avg[metric]
                if val > _metric['max_val']:
                    _metric['max_val'] = val
                if val < _metric['min_val']:
                    _metric['min_val'] = val
                count += 1
                metric_sum[metric] += val
                _metric['avg_val'] = metric_sum[metric] // count
            metrics_list = yield metrics_list

    def get_data(self, df, session=None):
        data = super(YandexHeatMapVisualization, self).get_data(df, session=session)
        data['legend'] = self.max_min_avg
        return data


class DeckScreengrid(BaseDeckGLViz):
    """deck.gl's ScreenGridLayer"""

    viz_type = 'deck_screengrid'
    verbose_name = _('Deck.gl - Screen Grid')
    spatial_control_keys = ['spatial']
    is_timeseries = True

    def query_obj(self):
        fd = self.form_data
        self.is_timeseries = fd.get('time_grain_sqla') or fd.get('granularity')
        return super(DeckScreengrid, self).query_obj()

    def get_properties(self, d):
        return {
            'position': d.get('spatial'),
            'weight': d.get(self.metric) or 1,
            '__timestamp': d.get(DTTM_ALIAS) or d.get('__time'),
        }


class DeckGrid(BaseDeckGLViz):
    """deck.gl's DeckLayer"""

    viz_type = 'deck_grid'
    verbose_name = _('Deck.gl - 3D Grid')
    spatial_control_keys = ['spatial']

    def get_properties(self, d):
        return {
            'position': d.get('spatial'),
            'weight': d.get(self.metric) or 1,
        }


class DeckPathViz(BaseDeckGLViz):
    """deck.gl's PathLayer"""

    viz_type = 'deck_path'
    verbose_name = _('Deck.gl - Paths')
    deck_viz_key = 'path'
    deser_map = {
        'json': json.loads,
        'polyline': polyline.decode,
    }

    def query_obj(self):
        d = super(DeckPathViz, self).query_obj()
        line_col = self.form_data.get('line_column')
        if d['metrics']:
            d['groupby'].append(line_col)
        else:
            d['columns'].append(line_col)
        return d

    def get_properties(self, d):
        fd = self.form_data
        deser = self.deser_map[fd.get('line_type')]
        path = deser(d[fd.get('line_column')])
        if fd.get('reverse_long_lat'):
            path = (path[1], path[0])
        return {
            self.deck_viz_key: path,
        }


class DeckPolygon(DeckPathViz):
    """deck.gl's Polygon Layer"""

    viz_type = 'deck_polygon'
    deck_viz_key = 'polygon'
    verbose_name = _('Deck.gl - Polygon')


class DeckHex(BaseDeckGLViz):
    """deck.gl's DeckLayer"""

    viz_type = 'deck_hex'
    verbose_name = _('Deck.gl - 3D HEX')
    spatial_control_keys = ['spatial']

    def get_properties(self, d):
        return {
            'position': d.get('spatial'),
            'weight': d.get(self.metric) or 1,
        }


class DeckGeoJson(BaseDeckGLViz):
    """deck.gl's GeoJSONLayer"""

    viz_type = 'deck_geojson'
    verbose_name = _('Deck.gl - GeoJSON')

    def query_obj(self):
        d = super(DeckGeoJson, self).query_obj()
        d['columns'] += [self.form_data.get('geojson')]
        d['metrics'] = []
        d['groupby'] = []
        return d

    def get_properties(self, d):
        geojson = d.get(self.form_data.get('geojson'))
        return json.loads(geojson)


class DeckArc(BaseDeckGLViz):
    """deck.gl's Arc Layer"""

    viz_type = 'deck_arc'
    verbose_name = _('Deck.gl - Arc')
    spatial_control_keys = ['start_spatial', 'end_spatial']

    def get_properties(self, d):
        return {
            'sourcePosition': d.get('start_spatial'),
            'targetPosition': d.get('end_spatial'),
        }

    def get_data(self, df, session=None):
        d = super(DeckArc, self).get_data(df, session=session)
        arcs = d['features']

        return {
            'arcs': arcs,
            'mapboxApiKey': config.get('MAPBOX_API_KEY'),
        }


class EventFlowViz(BaseViz):
    """A visualization to explore patterns in event sequences"""

    viz_type = 'event_flow'
    verbose_name = _('Event flow')
    credits = 'from <a href="https://github.com/williaster/data-ui">@data-ui</a>'
    is_timeseries = True

    def query_obj(self):
        query = super(EventFlowViz, self).query_obj()
        form_data = self.form_data

        event_key = form_data.get('all_columns_x')
        entity_key = form_data.get('entity')
        meta_keys = [
            col for col in form_data.get('all_columns')
            if col != event_key and col != entity_key
        ]

        query['columns'] = [event_key, entity_key] + meta_keys

        if form_data['order_by_entity']:
            query['orderby'] = [(entity_key, True)]

        return query

    def get_data(self, df, session=None):
        return df.to_dict(orient='records')


class PairedTTestViz(BaseViz):
    """A table displaying paired t-test values"""

    viz_type = 'paired_ttest'
    verbose_name = _('Time Series - Paired t-test')
    sort_series = False
    is_timeseries = True

    def get_data(self, df, session=None):
        """
        Transform received data frame into an object of the form:
        {
            'metric1': [
                {
                    groups: ('groupA', ... ),
                    values: [ {x, y}, ... ],
                }, ...
            ], ...
        }
        """
        fd = self.form_data
        groups = fd.get('groupby')
        metrics = fd.get('metrics')
        df.fillna(0)
        df = df.pivot_table(
            index=DTTM_ALIAS,
            columns=groups,
            values=metrics,
            margins_name=__('All')
        )
        cols = []
        # Be rid of falsey keys
        for col in df.columns:
            if col == '':
                cols.append('N/A')
            elif col is None:
                cols.append('NULL')
            else:
                cols.append(col)
        df.columns = cols
        data = {}
        series = df.to_dict('series')
        for nameSet in df.columns:
            # If no groups are defined, nameSet will be the metric name
            hasGroup = not isinstance(nameSet, string_types)
            Y = series[nameSet]
            d = {
                'group': nameSet[1:] if hasGroup else __('All'),
                'values': [
                    {'x': t, 'y': Y[t] if t in Y else None}
                    for t in df.index
                ],
            }
            key = nameSet[0] if hasGroup else nameSet
            if key in data:
                data[key].append(d)
            else:
                data[key] = [d]
        return data


class RoseViz(NVD3TimeSeriesViz):
    viz_type = 'rose'
    verbose_name = _('Time Series - Nightingale Rose Chart')
    sort_series = False
    is_timeseries = True

    def get_data(self, df, session=None):
        data = super(RoseViz, self).get_data(df, session=session)
        result = {}
        for datum in data:
            key = datum['key']
            for val in datum['values']:
                timestamp = val['x']
                if not isinstance(timestamp, str):
                    timestamp = str(val['x'])
                if not result.get(timestamp):
                    result[timestamp] = []
                value = 0 if math.isnan(val['y']) else val['y']
                result[timestamp].append({
                    'key': key,
                    'value': value,
                    'name': ', '.join(key) if isinstance(key, list) else key,
                    'time': str(val['x']),
                })
        return result


class PartitionViz(NVD3TimeSeriesViz):
    """
    A hierarchical data visualization with support for time series.
    """

    viz_type = 'partition'
    verbose_name = _('Partition Diagram')

    def query_obj(self):
        query_obj = super(PartitionViz, self).query_obj()
        time_op = self.form_data.get('time_series_option', 'not_time')
        # Return time series data if the user specifies so
        query_obj['is_timeseries'] = time_op != 'not_time'
        return query_obj

    def levels_for(self, time_op, groups, df):
        """
        Compute the partition at each `level` from the dataframe.
        """
        levels = {}
        for i in range(0, len(groups) + 1):
            agg_df = df.groupby(groups[:i]) if i else df
            levels[i] = (
                agg_df.mean() if time_op == 'agg_mean'
                else agg_df.sum(numeric_only=True))
        return levels

    def levels_for_diff(self, time_op, groups, df):
        # Obtain a unique list of the time grains
        times = list(set(df[DTTM_ALIAS]))
        times.sort()
        until = times[len(times) - 1]
        since = times[0]
        # Function describing how to calculate the difference
        func = {
            'point_diff': [
                pd.Series.sub,
                lambda a, b, fill_value: a - b,
            ],
            'point_factor': [
                pd.Series.div,
                lambda a, b, fill_value: a / float(b),
            ],
            'point_percent': [
                lambda a, b, fill_value=0: a.div(b, fill_value=fill_value) - 1,
                lambda a, b, fill_value: a / float(b) - 1,
            ],
        }[time_op]
        agg_df = df.groupby(DTTM_ALIAS).sum()
        levels = {0: pd.Series({
            m: func[1](agg_df[m][until], agg_df[m][since], 0)
            for m in agg_df.columns})}
        for i in range(1, len(groups) + 1):
            agg_df = df.groupby([DTTM_ALIAS] + groups[:i]).sum()
            levels[i] = pd.DataFrame({
                m: func[0](agg_df[m][until], agg_df[m][since], fill_value=0)
                for m in agg_df.columns})
        return levels

    def levels_for_time(self, groups, df):
        procs = {}
        for i in range(0, len(groups) + 1):
            self.form_data['groupby'] = groups[:i]
            df_drop = df.drop(groups[i:], 1)
            procs[i] = self.process_data(df_drop, aggregate=True).fillna(0)
        self.form_data['groupby'] = groups
        return procs

    def nest_values(self, levels, level=0, metric=None, dims=()):
        """
        Nest values at each level on the back-end with
        access and setting, instead of summing from the bottom.
        """
        if not level:
            return [{
                'name': m,
                'val': levels[0][m],
                'children': self.nest_values(levels, 1, m),
            } for m in levels[0].index]
        if level == 1:
            return [{
                'name': i,
                'val': levels[1][metric][i],
                'children': self.nest_values(levels, 2, metric, (i,)),
            } for i in levels[1][metric].index]
        if level >= len(levels):
            return []
        return [{
            'name': i,
            'val': levels[level][metric][dims][i],
            'children': self.nest_values(
                levels, level + 1, metric, dims + (i,),
            ),
        } for i in levels[level][metric][dims].index]

    def nest_procs(self, procs, level=-1, dims=(), time=None):
        if level == -1:
            return [{
                'name': m,
                'children': self.nest_procs(procs, 0, (m,)),
            } for m in procs[0].columns]
        if not level:
            return [{
                'name': t,
                'val': procs[0][dims[0]][t],
                'children': self.nest_procs(procs, 1, dims, t),
            } for t in procs[0].index]
        if level >= len(procs):
            return []
        return [{
            'name': i,
            'val': procs[level][dims][i][time],
            'children': self.nest_procs(procs, level + 1, dims + (i,), time),
        } for i in procs[level][dims].columns]

    def get_data(self, df, session=None):
        fd = self.form_data
        groups = fd.get('groupby', [])
        time_op = fd.get('time_series_option', 'not_time')
        if not len(groups):
            raise ValueError('Please choose at least one groupby')
        if time_op == 'not_time':
            levels = self.levels_for('agg_sum', groups, df)
        elif time_op in ['agg_sum', 'agg_mean']:
            levels = self.levels_for(time_op, groups, df)
        elif time_op in ['point_diff', 'point_factor', 'point_percent']:
            levels = self.levels_for_diff(time_op, groups, df)
        elif time_op == 'adv_anal':
            procs = self.levels_for_time(groups, df)
            return self.nest_procs(procs)
        else:
            levels = self.levels_for('agg_sum', [DTTM_ALIAS] + groups, df)
        return self.nest_values(levels)


class SpeedometerViz(BaseViz):
    viz_type = 'speedometer'
    verbose_name = _('Speedometer View')
    is_timeseries = False

    def query_obj(self):
        d = super(SpeedometerViz, self).query_obj()
        metric = self.form_data.get('metric')
        if not metric:
            raise Exception(_('Pick a metric!'))
        d['metrics'] = [self.form_data.get('metric')]
        self.form_data['metric'] = metric
        return d

    def get_data(self, df, session=None):
        form_data = self.form_data
        df.sort_values(by=df.columns[0], inplace=True)
        return {
            'data': df.values.tolist(),
            'subheader': form_data.get('subheader', ''),
        }


viz_types = {
    o.viz_type: o for o in globals().values()
    if (inspect.isclass(o) and issubclass(o, BaseViz) and o.viz_type not in config.get('VIZ_TYPE_BLACKLIST'))
}
